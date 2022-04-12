'''
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=1842931&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=1842931&score=0&sortType=5&page=1&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=1842931&score=0&sortType=5&page=2&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=1842931&score=0&sortType=5&page=3&pageSize=10&isShadowSku=0&rid=0&fold=1
'''

import requests
import json
import openpyxl as op
import pandas as pd
import jieba
from stylecloud import gen_stylecloud
from collections import Counter
from snownlp import SnowNLP
from pyecharts.charts import Bar
import pyecharts.options as opts
from pyecharts.charts import Line, Grid

def spider_jd():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='昵称')
    wb.cell(row=1, column=2, value='评分')
    wb.cell(row=1, column=3, value='产品类型')
    wb.cell(row=1, column=4, value='评论时间')
    wb.cell(row=1, column=5, value='评论点赞数')
    wb.cell(row=1, column=6, value='评论回复数')
    wb.cell(row=1, column=7, value='评论内容')
    count = 2

    for page in range(1, 100+1):

        print(f'-----------------正在抓取第{page}页数据中-----------------')
        url = f'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=1842931&score=0&sortType=5&page={page}&pageSize=10&isShadowSku=0&rid=0&fold=1'

        headers = {
            'Cookie': 'unpl=V2_ZzNtbUYHR0Z0W0RWKUsLV2JREA9LXhcQIFoVB30QC1JnUxUOclRCFnUUR1xnGF4UZwYZXEBcQBRFCEdkeBBVAWMDE1VGZxBFLV0CFSNGF1wjU00zQwBBQHcJFF0uSgwDYgcaDhFTQEJ2XBVQL0oMDDdRFAhyZ0AVRQhHZH4RWwdlBRtaQGdzEkU4dlF8GV4GYzMTbUNnAUEpDEJUehhbSGILFV9AUUoSdzhHZHg%3d; __jdv=76161171|baidu-pinzhuan|t_288551095_baidupinzhuan|cpc|0f3d30c8dba7459bb52f2eb5eba8ac7d_0_5a5b0b33ccfc4c3c88e4dcbb78ff1a6b|1630649001988; __jdu=496337863; areaId=27; PCSYCityID=CN_610000_0_0; shshshfpa=a3671771-e2dd-e5ca-3826-cc6c1253d728-1630649005; shshshfpb=zwlMB6I2rOyoYeuo0V6gtdw%3D%3D; user-key=8e0c5144-8b59-463a-82d8-113c57fd1256; _pst=18892080083_pie; unick=%E5%A4%A7%E5%B0%8F%E8%AF%B4%E5%AE%B63; pin=18892080083_pie; thor=F26D54FE61F2B579F14FE0C55C00221B9C479F0871905120BD991A994D4EA9C43A60A4EB5A55AA5EBC0029C3AE2909633951CEB21D44548EE3AD9276B1A650D1D5E938AB485D8708F1E767A8B44503EFA82588A0DE1C83E88CC8F22632DB9EBDB62F76AC5CAD097F2A1F3BAEF3AA4BC32C0956F76DEEA203421F2F95A71D634D205C08D6EE0E8BD6B0FBB114387CE7A39116936C0C32E4D40854ABA3D6780899; _tp=dr8%2BZKvIJSECFYFRRyrDdA%3D%3D; pinId=hX_o7kOWsfSxiEupljIIzQ; shshshfp=166a1d8dfb1b2b55419a4cb0f70894c0; __jda=122270672.496337863.1630649000.1630649000.1630649002.1; __jdc=122270672; token=9ab8d7eb0dbbf4d5c4ddb224b59b3c46,2,905916; __tk=VUoJVxeMWZjCWlfyV0kwVMbOnxjNiUVOixbzRUyDnxG,2,905916; shshshsID=1134730281e295e1469f554a1fe95c59_5_1630649056546; __jdb=122270672.7.496337863|1.1630649002; ipLoc-djd=27-2428-31523-56718; 3AB9D23F7A4B3C9B=IS7P7DMAVM4STNYR6HHL2B2DAETHTEWAQ22DWM4MKQ642L5GPJSZKE6HTB3IAS7O763NIN3SCCAV66ZY5JXZ2SCFV4; jwotest_product=99; JSESSIONID=8971DE7AF521E92FE4AE8A0C4518F6AA.s1',
            'Referer': 'https://item.jd.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'
        }

        # 获取响应
        resp = requests.get(url, headers=headers)
        json_data = json.loads(resp.text[20:-2])


        for cmts in json_data['comments']:

            # 昵称
            nickname = cmts['nickname']
            # 评分
            score = cmts['score']
            # 评论
            comments = cmts['content']
            # 产品类型
            product = cmts['productColor']

            # 评论时间
            time = cmts['referenceTime']

            # 评论点赞数
            starVote = cmts['usefulVoteCount']

            # 评论回复数
            cmtsReply = cmts['replyCount']

            wb.cell(row=count, column=1, value=nickname)
            wb.cell(row=count, column=2, value=score)
            wb.cell(row=count, column=3, value=product)
            wb.cell(row=count, column=4, value=time)
            wb.cell(row=count, column=5, value=starVote)
            wb.cell(row=count, column=6, value=cmtsReply)
            wb.cell(row=count, column=7, value=comments)


            print(nickname, score, comments, product, time, starVote, cmtsReply)
            count += 1

    ws.save('月饼.xlsx')

# 读取文件
rcv_data = pd.read_excel('./月饼.xlsx')
exist_col = rcv_data.dropna()  # 删除空行
print(exist_col.sample(5))
pd.set_option('display.max_columns', None)  # 显示完整的列
pd.set_option('display.max_rows', None)  # 显示完整的行
pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据
def visual_ciyun():
    # 读取文件
    rcv_data = pd.read_excel('./月饼.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    print(exist_col.sample(5))


    c_title = exist_col['评论内容'].tolist()

    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)

    # 设置停用词
    stop_words = []

    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-couch',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=stop_words
                   )
    print('词云图绘制成功！')

    # 词频设置
    all_words = [word for word in result.split(' ') if len(word) > 1 and word not in stop_words]
    wordcount = Counter(all_words).most_common(10)

    x_data, y_data = list(zip(*wordcount))
    print(x_data)
    print(y_data)

    c = (
        Bar()
            .add_xaxis(x_data)
            .add_yaxis("饼-词频展示图", y_data)
            .set_global_opts(title_opts=opts.TitleOpts(title="月饼-词频展示图"))
            .set_series_opts(
            label_opts=opts.LabelOpts(is_show=False),
            markpoint_opts=opts.MarkPointOpts(
                data=[
                    opts.MarkPointItem(type_="max", name="最大值"),
                    opts.MarkPointItem(type_="min", name="最小值"),
                    opts.MarkPointItem(type_="average", name="平均值"),
                ]
            ),
        )
            .render("数值展示.html")
    )

    '''c = (
        Bar()
            .add_xaxis(x_data)
            .add_yaxis("月饼-词频展示图", y_data)
            .set_global_opts(
            title_opts=opts.TitleOpts(title="月饼-词频展示图"),
            toolbox_opts=opts.ToolboxOpts(),
            legend_opts=opts.LegendOpts(is_show=False),
        )
            .render("月饼-词频展示图x.html")
    )
    print('月饼-词频展示图绘制成功！')'''


def product_visual():

    product = rcv_data['产品类型'].value_counts()
    x_data = product.index.tolist()
    y_data = product.tolist()

    print(x_data)
    print(y_data)

    c = (
        Bar()
            .add_xaxis(x_data)
            .add_yaxis("月饼口味", y_data)
            .reversal_axis()
            .set_series_opts(label_opts=opts.LabelOpts(position="right"))
            .set_global_opts(title_opts=opts.TitleOpts(title="月饼类型销售饼图"))
            .render("月饼类型销售饼图.html")
    )



    '''data_pair = [list(z) for z in zip(x_data, y_data)]
    data_pair.sort(key=lambda x: x[1])

    (
        Pie(init_opts=opts.InitOpts(width="1600px", height="800px", bg_color="#2c343c"))
            .add(
            series_name="访问来源",
            data_pair=data_pair,
            rosetype="radius",
            radius="55%",
            center=["50%", "50%"],
            label_opts=opts.LabelOpts(is_show=False, position="center"),
        )
            .set_global_opts(
            title_opts=opts.TitleOpts(
                title="月饼类型销售-图",
                pos_left="center",
                pos_top="20",
                title_textstyle_opts=opts.TextStyleOpts(color="#fff"),
            ),
            legend_opts=opts.LegendOpts(is_show=False),
        )
            .set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{a} <br/>{b}: {c} ({d}%)"
            ),
            label_opts=opts.LabelOpts(color="rgba(255, 255, 255, 0.3)"),
        )
            .render("月饼类型销售饼图.html")
    )'''


def max_data():
    # 读取数据
    pd_data = pd.read_excel('./月饼.xlsx')
    pd.set_option('display.max_columns', None)  # 显示完整的列
    pd.set_option('display.max_rows', None)  # 显示完整的行
    pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据

    # 最多点赞数
    max_star = pd_data[pd_data['评论点赞数'] == pd_data['评论点赞数'].max()]
    print(max_star)

    # 最多回复数
    max_reply = pd_data[pd_data['评论点赞数'] == pd_data['评论点赞数'].max()]
    print(max_reply)

    # 情感分析
    pinglun = u'''
        连续20年香港销售冠军?，没有月饼能在中秋打败它
        首创流心奶黄，是开山鼻祖哦
        100%香港制作，品质数年如一日，稳定的很
        打开流心奶黄月饼礼盒，外观是淡淡的乳黄色，上面印着烫金字“美心流心奶黄”。里面有8枚流心奶黄月饼，一个是45克，个头刚刚好。这种大小就比较适合小仙女和长辈们吃，一个吃下去刚刚好～
        流心月饼最好的吃法，就是微波炉加热5-7秒，亲测7秒最佳，如果放在冰箱里冰过，就热10秒。
        拿出来后，饼皮热乎乎的，软软的，掰开来后，奶黄慢慢流淌出来，闻着好香好诱人！
    '''

    Sentiment_analysis = SnowNLP(pinglun).sentiments
    print(Sentiment_analysis)

def extr_time():
    # 读取数据
    pd_data = pd.read_excel('./月饼.xlsx')

    # 年份及其数量
    day = pd.DatetimeIndex(pd_data['评论时间']).day.value_counts()
    x_day =day.index.tolist()
    y_day = day.tolist()

    print(x_day)
    print(y_day)

    c = (
        Line()
            .add_xaxis(xaxis_data=['16', '17', '6', '9', '4', '3', '22', '2,' '8', '26', '14', '7', '23', '5', '12', '24', '10', '18', '20', '1', '19', '13', '15', '11', '30', '27', '29', '28', '31', '21', '25'])
            .add_yaxis(
            "日-销量",
            y_axis=y_day,
            linestyle_opts=opts.LineStyleOpts(width=2),
        )
            .set_global_opts(
            title_opts=opts.TitleOpts(title="日-月饼销量折线图"),
            xaxis_opts=opts.AxisOpts(name="x"),
            yaxis_opts=opts.AxisOpts(
                type_="log",
                name="y",
                splitline_opts=opts.SplitLineOpts(is_show=True),
                is_scale=True,
            ),
        )
            .render("line_yaxis_log.html")
    )



if __name__ == '__main__':
    spider_jd()
    #visual_ciyun()
    #product_visual()
    #max_data()
    #extr_time()