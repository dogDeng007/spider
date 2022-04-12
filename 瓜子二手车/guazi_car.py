'''

瓜子二手车

'''

import requests
from lxml import etree
from fake_useragent import UserAgent
import openpyxl as op
import time, random


def spider_car():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='车辆概况')
    wb.cell(row=1, column=2, value='车龄')
    wb.cell(row=1, column=3, value='全款价')
    wb.cell(row=1, column=4, value='汽车原价')
    wb.cell(row=1, column=5, value='汽车排量')
    wb.cell(row=1, column=6, value='过户情况')
    wb.cell(row=1, column=7, value='变速箱')
    wb.cell(row=1, column=8, value='表显里程')
    wb.cell(row=1, column=9, value='最低首付')
    count = 2

    for page in range(1, 50+1):
        print(f'-----------正在采集第{page}页数据-----------')
        url = f'https://www.guazi.com/cd/buy/o{page}/#bread'

        headers = {
            'Referer': 'https://www.guazi.com/xa/buy/o1/',
            'Cookie': 'uuid=aecbdf7d-8648-4f28-dbf9-902293bbf045; clueSourceCode=%2A%2300; user_city_id=176; ganji_uuid=6311203881588826872627; guazitrackersessioncadata=%7B%22ca_kw%22%3A%22%25e7%2593%259c%25e5%25ad%2590%25e4%25ba%258c%25e6%2589%258b%25e8%25bd%25a6%22%7D; sessionid=d173d24e-e570-4019-cc19-47bcd5e33348; lg=1; Hm_lvt_bf3ee5b290ce731c7a4ce7a617256354=1623491103; close_finance_popup=2021-06-12; cainfo=%7B%22ca_a%22%3A%22-%22%2C%22ca_b%22%3A%22-%22%2C%22ca_s%22%3A%22pz_baidu%22%2C%22ca_n%22%3A%22pcbiaoti%22%2C%22ca_medium%22%3A%22-%22%2C%22ca_term%22%3A%22-%22%2C%22ca_content%22%3A%22-%22%2C%22ca_campaign%22%3A%22-%22%2C%22ca_kw%22%3A%22%25e7%2593%259c%25e5%25ad%2590%25e4%25ba%258c%25e6%2589%258b%25e8%25bd%25a6%22%2C%22ca_i%22%3A%22-%22%2C%22scode%22%3A%22-%22%2C%22keyword%22%3A%22-%22%2C%22ca_keywordid%22%3A%22-%22%2C%22display_finance_flag%22%3A%22-%22%2C%22platform%22%3A%221%22%2C%22version%22%3A1%2C%22client_ab%22%3A%22-%22%2C%22guid%22%3A%22aecbdf7d-8648-4f28-dbf9-902293bbf045%22%2C%22ca_city%22%3A%22xa%22%2C%22sessionid%22%3A%22d173d24e-e570-4019-cc19-47bcd5e33348%22%7D; _gl_tracker=%7B%22ca_source%22%3A%22-%22%2C%22ca_name%22%3A%22-%22%2C%22ca_kw%22%3A%22-%22%2C%22ca_id%22%3A%22-%22%2C%22ca_s%22%3A%22self%22%2C%22ca_n%22%3A%22-%22%2C%22ca_i%22%3A%22-%22%2C%22sid%22%3A64163631266%7D; cityDomain=xa; preTime=%7B%22last%22%3A1623491941%2C%22this%22%3A1623491102%2C%22pre%22%3A1623491102%7D; Hm_lpvt_bf3ee5b290ce731c7a4ce7a617256354=1623491941',
            'User-Agent': str(UserAgent().random)
        }
        resp = requests.get(url, headers = headers).text

        html = etree.HTML(resp)
        lis = html.xpath("//ul[@class='carlist clearfix js-top']/li")

        for li in lis:

            car_info = li.xpath('./a/h2/text()')
            car_info = ''.join(car_info)                    # 汽车概况

            car_age = li.xpath('./a/div[1]/text()[1]')      # 车龄
            car_age = ''.join(car_age)

            car_money = li.xpath('./a/div[2]/p/text()')
            car_money = (''.join(car_money)).strip()+'万'        # 全款价

            car_money_before = li.xpath('./a/div[2]/em/text()')
            car_money_before = ''.join(car_money_before)        # 汽车原价

            # 下一页数据连接
            links = li.xpath('./a/@href')
            x = 'https://www.guazi.com'
            page_down = [x + i for i in links]

            for page in page_down:
                response = requests.get(page, headers = headers)
                if response.status_code == 200:
                    html = etree.HTML(response.text)

                    car_displace = html.xpath("//li[@class='three']/span/text()")
                    car_displace = ''.join(car_displace)                            # 汽车排量

                    car_transf = html.xpath("//ul[@class='basic-eleven clearfix']/li[@class='seven']/div/text()")
                    car_transf = ''.join(car_transf).strip()                                      # 过户情况

                    car_change = html.xpath("//ul[@class='assort clearfix']/li[@class='last']/span/text()")
                    car_change = ''.join(car_change)                                       # 变速箱

                    car_mile = html.xpath("//li[@class='two']/span/text()")
                    car_mile = ''.join(car_mile)                                            # 表显里程

                    car_pay = html.xpath("//a[@class='loanbox js-loan']/span[@class='f24']/text()")
                    car_pay = ''.join(car_pay)
                    if car_pay == '':
                        car_pay = '该车暂不支持分期'                                        # 最低首付

                    wb.cell(row=count, column=1, value=car_info)
                    wb.cell(row=count, column=2, value=car_age)
                    wb.cell(row=count, column=3, value=car_money)
                    wb.cell(row=count, column=4, value=car_money_before)
                    wb.cell(row=count, column=5, value=car_displace)
                    wb.cell(row=count, column=6, value=car_transf)
                    wb.cell(row=count, column=7, value=car_change)
                    wb.cell(row=count, column=8, value=car_mile)
                    wb.cell(row=count, column=9, value=car_pay)
                    print('-----正在采集第' + str(count - 1) + '条数据-----')
                    print(car_info, car_age, car_money, car_money_before, car_displace, car_transf, car_change, car_mile, car_pay)
                    count += 1
                time.sleep(random.random() * 3)
                ws.save('瓜子二手车.xlsx')

if __name__ == '__main__':
    spider_car()



