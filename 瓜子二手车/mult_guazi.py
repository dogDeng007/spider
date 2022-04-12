import requests
from queue import Queue
import threading
from fake_useragent import UserAgent
from lxml import etree
import openpyxl as op

from pyecharts.charts import Bar, Pie, Funnel, Page  # 各个图形的类
from pyecharts.charts import Line
import pandas as pd
from pyecharts import options as opts

class Procuder(threading.Thread):

    headers = {
        'Referer': 'https://www.guazi.com/xa/buy/o1/',
        'Cookie': 'uuid=aecbdf7d-8648-4f28-dbf9-902293bbf045; clueSourceCode=%2A%2300; user_city_id=176; ganji_uuid=6311203881588826872627; guazitrackersessioncadata=%7B%22ca_kw%22%3A%22%25e7%2593%259c%25e5%25ad%2590%25e4%25ba%258c%25e6%2589%258b%25e8%25bd%25a6%22%7D; sessionid=d173d24e-e570-4019-cc19-47bcd5e33348; lg=1; Hm_lvt_bf3ee5b290ce731c7a4ce7a617256354=1623491103; close_finance_popup=2021-06-12; cainfo=%7B%22ca_a%22%3A%22-%22%2C%22ca_b%22%3A%22-%22%2C%22ca_s%22%3A%22pz_baidu%22%2C%22ca_n%22%3A%22pcbiaoti%22%2C%22ca_medium%22%3A%22-%22%2C%22ca_term%22%3A%22-%22%2C%22ca_content%22%3A%22-%22%2C%22ca_campaign%22%3A%22-%22%2C%22ca_kw%22%3A%22%25e7%2593%259c%25e5%25ad%2590%25e4%25ba%258c%25e6%2589%258b%25e8%25bd%25a6%22%2C%22ca_i%22%3A%22-%22%2C%22scode%22%3A%22-%22%2C%22keyword%22%3A%22-%22%2C%22ca_keywordid%22%3A%22-%22%2C%22display_finance_flag%22%3A%22-%22%2C%22platform%22%3A%221%22%2C%22version%22%3A1%2C%22client_ab%22%3A%22-%22%2C%22guid%22%3A%22aecbdf7d-8648-4f28-dbf9-902293bbf045%22%2C%22ca_city%22%3A%22xa%22%2C%22sessionid%22%3A%22d173d24e-e570-4019-cc19-47bcd5e33348%22%7D; _gl_tracker=%7B%22ca_source%22%3A%22-%22%2C%22ca_name%22%3A%22-%22%2C%22ca_kw%22%3A%22-%22%2C%22ca_id%22%3A%22-%22%2C%22ca_s%22%3A%22self%22%2C%22ca_n%22%3A%22-%22%2C%22ca_i%22%3A%22-%22%2C%22sid%22%3A64163631266%7D; cityDomain=xa; preTime=%7B%22last%22%3A1623491941%2C%22this%22%3A1623491102%2C%22pre%22%3A1623491102%7D; Hm_lpvt_bf3ee5b290ce731c7a4ce7a617256354=1623491941',
        'User-Agent': str(UserAgent().random)
    }

    def __init__(self, page_queue, img_queue, *args, **kwargs):
        super(Procuder, self).__init__(*args, **kwargs)
        self.page_queue = page_queue
        self.img_queue = img_queue


    def run(self) -> None:
        while True:
            if self.page_queue.empty():
                break
            url = self.page_queue.get()
            self.parse_page(url)


    def parse_page(self, url):
        resp = requests.get(url, headers=self.headers).text

        html = etree.HTML(resp)
        lis = html.xpath("//ul[@class='carlist clearfix js-top']/li")

        for li in lis:

            car_info = li.xpath('./a/h2/text()')
            car_info = ''.join(car_info)  # 汽车概况

            car_age = li.xpath('./a/div[1]/text()[1]')  # 车龄
            car_age = ''.join(car_age)

            car_money = li.xpath('./a/div[2]/p/text()')
            car_money = (''.join(car_money)).strip()   # 全款价

            car_money_before = li.xpath('./a/div[2]/em/text()')
            car_money_before = ''.join(car_money_before)  # 汽车原价

            # 下一页数据连接
            links = li.xpath('./a/@href')
            x = 'https://www.guazi.com'
            page_down = [x + i for i in links]

            for page in page_down:
                response = requests.get(page, headers=self.headers)
                if response.status_code == 200:
                    html = etree.HTML(response.text)

                    car_displace = html.xpath("//li[@class='three']/span/text()")
                    car_displace = ''.join(car_displace)  # 汽车排量

                    car_transf = html.xpath("//ul[@class='basic-eleven clearfix']/li[@class='seven']/div/text()")
                    car_transf = ''.join(car_transf).strip()  # 过户情况

                    car_change = html.xpath("//ul[@class='assort clearfix']/li[@class='last']/span/text()")
                    car_change = ''.join(car_change)  # 变速箱

                    car_mile = html.xpath("//li[@class='two']/span/text()")
                    car_mile = ''.join(car_mile)  # 表显里程

                    car_pay = html.xpath("//a[@class='loanbox js-loan']/span[@class='f24']/text()")
                    car_pay = ''.join(car_pay)
                    self.img_queue.put((car_info, car_age, car_money, car_money_before, car_displace, car_transf, car_change, car_mile, car_pay))

class Consumer(threading.Thread):

    def __init__(self, page_queue, img_queue, *args, **kwargs):
        super(Consumer, self).__init__(*args, **kwargs)
        self.page_queue = page_queue
        self.img_queue = img_queue


    def run(self) -> None:
        ws = op.Workbook()
        wb = ws.create_sheet(index=0)

        wb.cell(row=1, column=1, value='车辆概况')
        wb.cell(row=1, column=2, value='车龄')
        wb.cell(row=1, column=3, value='全款价')
        wb.cell(row=1, column=4, value='汽车原价')
        wb.cell(row=1, column=5, value='汽车排量')
        wb.cell(row=1, column=6, value='过户情况')
        wb.cell(row=1, column=7, value='变速箱')
        wb.cell(row=1, column=8, value='表显里程')
        wb.cell(row=1, column=9, value='最低首付')
        count = 2

        while True:
            if self.img_queue.empty() and self.page_queue.empty():
                break
            car_info, car_age, car_money, car_money_before, car_displace, car_transf, car_change, car_mile, car_pay = self.img_queue.get()

            wb.cell(row=count, column=1, value=car_info)
            wb.cell(row=count, column=2, value=car_age)
            wb.cell(row=count, column=3, value=car_money)
            wb.cell(row=count, column=4, value=car_money_before)
            wb.cell(row=count, column=5, value=car_displace)
            wb.cell(row=count, column=6, value=car_transf)
            wb.cell(row=count, column=7, value=car_change)
            wb.cell(row=count, column=8, value=car_mile)
            wb.cell(row=count, column=9, value=car_pay)

            print(car_info, car_age, car_money, car_money_before, car_displace, car_transf, car_change, car_mile,
                  car_pay)
            count += 1
            ws.save('多线程瓜子二手车.xlsx')

def main():
    page_queue = Queue(100)
    img_queue = Queue(1000)

    for page in range(1, 50+1):
        url = f'https://www.guazi.com/cd/buy/o{page}/#bread'
        page_queue.put(url)

    # 开启5个生产者
    for x in range(5):
        t = Procuder(page_queue, img_queue)
        t.start()

    # 开启10个消费者
    for x in range(10):
        t = Consumer(page_queue, img_queue)
        t.start()
        t.join()


pd_data = pd.read_excel('多线程瓜子二手车.xlsx')
# 汽车名称分析
def name_analysis():
    # 提取汽车名称
    pd_data['汽车名称'] = pd_data['车辆概况'].map(lambda x: x.split(" ")[0])

    name = pd_data['汽车名称'].value_counts()
    # 统计汽车名称(前20)
    name1 = name.index.tolist()[:20]
    # 车龄汽车名称对应数量(前20)
    name2 = name.tolist()[:20]
    return name1, name2


# 表显里程分析
def mile_analysis():
    pd_data.loc[:, '表显里程1'] = pd_data['表显里程'].str.replace('万公里', '').astype('float32')  # 去除 30 ’万公里‘
    pd_data['里程区间'] = pd.cut(pd_data['表显里程1'], [0, 2, 4, 6, 8, 10, 20],
                             labels=['0-2', '2-4', '4-6', '6-8', '8-10', '>10'])
    mile = pd_data['里程区间'].value_counts()
    mile1 = mile.index.tolist()  # 变速箱种类
    mile2 = mile.tolist()  # 变速箱种类对应数量'''
    return mile1, mile2


# 最低首付分析
def min_pay():
    pd_data = pd.read_excel('多线程瓜子二手车777.xlsx')
    # 删除空格
    pd_data = pd_data.dropna(subset=['最低首付'])
    # 划分价格区间
    pd_data['价格区间'] = pd.cut(pd_data['最低首付'], [0, 3, 5, 8, 10, 15], labels=['0-3', '3-5', '5-8', '8-10', '>10'])
    # 统计数量
    price = pd_data['价格区间'].value_counts()
    price1 = price.index.tolist()  # 变速箱种类
    price2 = price.tolist()  # 变速箱种类对应数量'''

    return price1, price2



# 过户情况分析
def trans_cars():
    tras = pd_data['过户情况'].value_counts()
    tras1 = tras.index.tolist()  # 变速箱种类
    tras2 = tras.tolist()  # 变速箱种类对应数量

    return tras1, tras2


# 汽车排量分析
def disp_cars():
    disp = pd_data['汽车排量'].value_counts()
    disp1 = disp.index.tolist()  # 变速箱种类
    disp2 = disp.tolist()  # 变速箱种类对应数量
    print(disp1)
    print(disp2)

    return disp1, disp2

# 变速箱分析
def chag_cars():
    chg = pd_data['变速箱'].value_counts()
    chg1 = chg.index.tolist()  # 变速箱种类
    chg2 = chg.tolist()  # 变速箱种类对应数量

    return chg1, chg2


# 车龄分析
def ages_car():
    age = pd_data['车龄'].value_counts()
    age1 = age.index.tolist()  # 车龄年代分布
    age2 = age.tolist()  # 车龄年代分布数量

    return age1, age2


    pd_data.to_excel('多线程瓜子二手车777.xlsx')


# 画图
def draw_cars():
    x_data, y_data = ages_car()

    # 1、柱状图
    def barPage() -> Bar:
        bar = (
            Bar()
                .add_xaxis(x_data)
                .add_yaxis("柱状图", y_data)
                .set_global_opts(
                title_opts=opts.TitleOpts(title="柱状图分析"),
                legend_opts=opts.LegendOpts(is_show=False), )
        )
        return bar

    # 2、饼图
    def piePage() -> Pie:
        pie = (
            Pie()
                .add("", [list(z) for z in zip(x_data, y_data)])
                .set_global_opts(title_opts=opts.TitleOpts(title="饼图分析"))
                .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}"))
        )
        return pie

    # 3、折线图
    def linePage() -> Line:
        line = (
            Line()
                .add_xaxis(x_data)
                .add_yaxis("折线图", y_data)
                .set_global_opts(title_opts=opts.TitleOpts(title="折线图分析"))
        )
        return line

    # 4、漏斗图
    def funnelPage() -> Funnel:
        funnel = (
            Funnel()
                .add("漏斗图", [list(z) for z in zip(x_data, y_data)])
                .set_global_opts(title_opts=opts.TitleOpts(title="漏斗图分析"))
        )
        return funnel


    # 上面是6个图形的代码，下面利用Page进行组合
    # !!! 关键步骤
    page = (
        Page(layout=Page.DraggablePageLayout)
            .add(
            barPage(),
            piePage(),
            linePage(),
            funnelPage())
    )

    page.render("车龄信息展示.html")
    print('绘图完成！')



if __name__ == '__main__':
    main()