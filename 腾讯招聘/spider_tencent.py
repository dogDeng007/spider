'''

https://careers.tencent.com/tencentcareer/api/post/Query?timestamp=1623410681974&countryId=&cityId=&bgIds=&productId=&categoryId=&parentCategoryId=&attrId=&keyword=&pageIndex=2&pageSize=10&language=zh-cn&area=cn
https://careers.tencent.com/tencentcareer/api/post/Query?timestamp=1623410723340&countryId=&cityId=&bgIds=&productId=&categoryId=&parentCategoryId=&attrId=&keyword=&pageIndex=3&pageSize=10&language=zh-cn&area=cn
https://careers.tencent.com/tencentcareer/api/post/Query?timestamp=1623410736923&countryId=&cityId=&bgIds=&productId=&categoryId=&parentCategoryId=&attrId=&keyword=&pageIndex=4&pageSize=10&language=zh-cn&area=cn
'''

import requests
from fake_useragent import UserAgent
import openpyxl as op
import pandas as pd
import time, random
import jieba
from stylecloud import gen_stylecloud
from pyecharts.charts import Bar,Pie
from pyecharts import options as opts

from pyecharts.charts import Map  # 导入模

# 爬取数据
def spider_tx():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='职位名称')
    wb.cell(row=1, column=2, value='国家')
    wb.cell(row=1, column=3, value='城市')
    wb.cell(row=1, column=4, value='职位分类')
    wb.cell(row=1, column=5, value='职位更新时间')
    wb.cell(row=1, column=6, value='职位要求')
    count = 2

    for page in range(1, 200+1):
        print(f'------------------正在爬取第{page}页数据------------------')
        url = f'https://careers.tencent.com/tencentcareer/api/post/Query?timestamp=1623410681974&countryId=&cityId=&bgIds=&productId=&categoryId=&parentCategoryId=&attrId=&keyword=&pageIndex={page}&pageSize=10&language=zh-cn&area=cn'

        headers= {
            'referer': f'https://careers.tencent.com/search.html?index={page}',
            'user-agent': str(UserAgent().random)
        }
        resp = requests.get(url, headers = headers)

        # 获取所有的职位信息
        jobs = resp.json()['Data']['Posts']

        # 获取每一个岗位的内部信息
        for job in jobs:
            post_name = job['RecruitPostName']  # 职位名称
            country_name = job['CountryName']  # 国家
            loc_name = job['LocationName']  # 城市
            category_name = job['CategoryName']  # 职位分类
            last_up_time = job['LastUpdateTime']  # 职位更新时间
            responsibility = job['Responsibility']  # 职位要求

            print(post_name, country_name, loc_name, category_name, last_up_time, responsibility)

            wb.cell(row=count, column=1, value=post_name)
            wb.cell(row=count, column=2, value=country_name)
            wb.cell(row=count, column=3, value=loc_name)
            wb.cell(row=count, column=4, value=category_name)
            wb.cell(row=count, column=5, value=last_up_time)
            wb.cell(row=count, column=6, value=responsibility)

            print('-----正在采集第' + str(count - 1) + '条数据-----')
            count += 1
        time.sleep(random.random() * 3)
    ws.save('腾讯职位.xlsx')

#绘制词云
def ciyun():
    rcv_data = pd.read_excel('腾讯职位.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    c_title = exist_col['职位要求'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fab fa-qq',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=['和', '与', '的', '及', '等', '对', '并', '', '', '', '', '', '', '']
                   )
    print('绘图成功！')


def pietu():
    pd_data = pd.read_excel('腾讯职位.xlsx')
    x = pd_data['职位分类'].value_counts()
    x1_data = x.index.tolist() #index, index.values, index.tolist()
    x2_data = x.tolist()
    print(x1_data)
    print(x2_data)

    y = pd_data['城市'].value_counts()
    y1_data = y.index.tolist()
    y2_data = y.tolist()
    print(y1_data)
    print(y2_data)



    # 绘制区域饼图
    c = (
        Pie()
            .add("", [list(z) for z in zip(x1_data, x2_data)])    # zip函数两个部分组合在一起list(zip(x,y))-----> [(x,y)]
            .set_global_opts(title_opts=opts.TitleOpts(title="Pie-腾讯招聘职位分类饼图"))     # 标题
            .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}: {d}"))      # 数据标签设置
    )
    c.render('腾讯招聘职位分类饼图.shtml')

    # 绘制区域柱状图
    bar = (
        Bar()
            .add_xaxis(x1_data)
            .add_yaxis("腾讯招聘职位分类柱状图", x2_data)
            .set_global_opts(
            title_opts=opts.TitleOpts(title="Bar-腾讯招聘职位分类柱状图"),
            legend_opts=opts.LegendOpts(is_show=False),
        )
    )
    bar.render('腾讯招聘职位分类柱状图.shtml')

    map = (
        Map()
            .add("腾讯职位分布图", [list(z) for z in zip(y1_data, y2_data)], "china")
            .set_global_opts(
            title_opts=opts.TitleOpts(title="腾讯职位分布图"),
            visualmap_opts=opts.VisualMapOpts(max_=1100, is_piecewise=True),
        )
    )
    map.render('腾讯职位分布图.shtml')


if __name__ == '__main__':
    #spider_tx()
    ciyun()
    #pietu()
