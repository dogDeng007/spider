'''
https://www.huya.com/cache.php?m=LiveList&do=getLiveListByPage&gameId=2168&tagAll=0&callback=getLiveListJsonpCallback&page=1
https://www.huya.com/cache.php?m=LiveList&do=getLiveListByPage&gameId=2168&tagAll=0&callback=getLiveListJsonpCallback&page=2
'''
import time

import requests
from fake_useragent import UserAgent
import json
import openpyxl as op
import os
import random
import numpy as np
import PIL.Image as Image
import shutil


def spider_huya():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='房间名称')
    wb.cell(row=1, column=2, value='封面照片')
    wb.cell(row=1, column=3, value='昵称')
    wb.cell(row=1, column=4, value='头像')
    wb.cell(row=1, column=5, value='直播间当前人数')
    count = 2

    # 爬取图片的list
    save_pic = []
    for page in range(1, 6+1):
        print(f'--------------正在打印第{page}页小姐姐信息--------------')
        url = f'https://www.huya.com/cache.php?m=LiveList&do=getLiveListByPage&gameId=2168&tagAll=0&callback=getLiveListJsonpCallback&page={page}'

        headers = {
            'cookie': '__yamid_tt1=0.17768755672559844; __yamid_new=C9662ED452B00001997340851CC8140B; game_did=j7Os0i0Txedw1cLUtP0vmKnalM1x65kO3rE; SoundValue=0.50; alphaValue=0.80; guid=0a42cb71a121c360e701bcfbbdfb20c9; udb_guiddata=b45f59af594a4e878296ed13cde65858; udb_anouid=1461170529732; isInLiveRoom=true; Hm_lvt_51700b6c722f5bb4cf39906a596ea41f=1623400798,1624888823,1624928362; udb_passdata=3; __yasmid=0.17768755672559844; _yasids=__rootsid%3DC96BDFA246600001458617807F641C12; Hm_lpvt_51700b6c722f5bb4cf39906a596ea41f=1624928418; huya_web_rep_cnt=137',
            'referer': 'https://www.huya.com/g/2168',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.8 Safari/537.36'
        }
        try:
            resp = requests.get(url, headers = headers)
        except:
            print('something wrong!')

        text = resp.text[25:-1]            # 去除text格式前后无效字符
        json_text = json.loads(text)       # 转为标准json格式数据集
        cmts_list = json_text['data']['datas']      # 获取详细信息


        for item in cmts_list:
            room_name = item['roomName']        # 房间名称
            cover_link = item['screenshot']     # 封面照片
            nick = item['nick']                 # 昵称
            prof_phot = item['avatar180']       # 头像
            view_num = item['totalCount']       # 直播间当前人数

            wb.cell(row=count, column=1, value=room_name)
            wb.cell(row=count, column=2, value=cover_link)
            wb.cell(row=count, column=3, value=nick)
            wb.cell(row=count, column=4, value=prof_phot)
            wb.cell(row=count, column=5, value=view_num)
            print('cover_link', cover_link)
            print('nick', type(nick))
            save_pic.append((cover_link, nick))
            print(save_pic)

            print(f'-----正在采集第{page}页的第{count}条数据-----')
            print(room_name, cover_link, nick, prof_phot, view_num)
            count += 1

    time.sleep(random.random()*5)
    ws.save('虎牙小姐姐1.xlsx')
    return save_pic


# 图片保存路径
pic_path = './pic'
def down_huya(save_pic):

    # 判断文件夹是否存在
    if not os.path.exists(pic_path):
        os.mkdir(pic_path)
    else:
        # 判断文件夹是否为空
        shutil.rmtree(pic_path)
        os.mkdir(pic_path)

    count = 0

    for num, (pic_img, pic_name) in enumerate(save_pic):
        r = requests.get(pic_img)
        pic = r.content
        try:
            with open('./pic/{}.jpg'.format(pic_name), 'wb') as fin:
                print(f'正在爬取第{count}张图片')
                fin.write(pic)
                print('{}.jpg----下载成功'.format(pic_name))
        except:
            print('下载失败！')
        count += 1

def draw_girls():
    save_path = r'./pic/'
    # 设置心性图片矩阵
    HEART = [[0, 0, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0],
             [0, 1, 1, 1, 1, 0, 0, 1, 1, 1, 1, 0],
             [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
             [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
             [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
             [0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0],
             [0, 0, 1, 1, 1, 1, 1, 1, 1, 1, 0, 0],
             [0, 0, 0, 1, 1, 1, 1, 1, 1, 0, 0, 0],
             [0, 0, 0, 0, 1, 1, 1, 1, 0, 0, 0, 0],
             [0, 0, 0, 0, 0, 1, 1, 0, 0, 0, 0, 0]]

    # 定义相关参数
    SIZE = 100  # 每张图片的尺寸，越大越清晰
    N = 1  # 每个点位上放置1*1张图片

    # 计算相关参数
    width = np.shape(HEART)[1] * N * SIZE  # 照片墙宽度
    height = np.shape(HEART)[0] * N * SIZE  # 照片墙高度
    n_img = np.sum(HEART) * (N ** 2)  # 照片墙需要的照片数
    filenames = random.sample(os.listdir(save_path), n_img)  # 随机选取n_img张照片
    filenames = [save_path + f for f in filenames]

    print('宝宝开始集合！')
    # 绘制爱心墙
    img_bg = Image.new('RGB', (width, height))  # 设置照片墙背景
    i = 0
    for y in range(np.shape(HEART)[0]):
        for x in range(np.shape(HEART)[1]):
            if HEART[y][x] == 1:  # 如果需要填充
                pos_x = x * N * SIZE  # 填充起始X坐标位置
                pos_y = y * N * SIZE  # 填充起始Y坐标位置
                for yy in range(N):
                    for xx in range(N):
                        img = Image.open(filenames[i])
                        img = img.resize((SIZE, SIZE), Image.ANTIALIAS)
                        img_bg.paste(img, (pos_x + xx * SIZE, pos_y + yy * SIZE))
                        i += 1
    # 保存图片
    img_bg.save('love.jpg')
    print('宝宝集合完毕！')


if __name__ == '__main__':

    down_huya(spider_huya())
    spider_huya()
    draw_girls()
