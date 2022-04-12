from fake_useragent import UserAgent
import requests
from lxml import etree
import openpyxl as op
import pandas as pd
import jieba
from stylecloud import gen_stylecloud
from pyecharts.charts import Funnel, Page  # 各个图形的类
from pyecharts.charts import Line
import pandas as pd
import jieba
from stylecloud import gen_stylecloud
from pyecharts.charts import Bar,Pie
from pyecharts import options as opts

from pyecharts import options as opts
from pyecharts.charts import Bar
from pyecharts.commons.utils import JsCode
from pyecharts.faker import Faker

# 抓取页面
def spider_page():
    # 创建workbook
    ws = op.Workbook()
    # 创建worksheet
    wb = ws.create_sheet(index=0)

    # 创建表头
    wb.cell(row=1, column=1, value='作者')
    wb.cell(row=1, column=2, value='时间')
    wb.cell(row=1, column=3, value='星评')
    wb.cell(row=1, column=4, value='评分')
    wb.cell(row=1, column=5, value='赞同')
    wb.cell(row=1, column=6, value='评论')
    count = 2

    # 多页抓取
    for page in range(1, 25 + 1):
        print(f'-----------------正在抓取第{page}页影评-----------------')
        url = f'https://movie.douban.com/subject/35202793/comments?percent_type=&start={(page - 1) * 20}&limit=20&status=P&sort=new_score&comments_only=1'
        print(url)

        headers = {
            'Cookie': 'bid=Hgm5jLcvu9M; dbcl2="153819075:MtvprK7V/38"; ck=8_y9; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1629936489%2C%22https%3A%2F%2Fopen.weixin.qq.com%2F%22%5D; _pk_ses.100001.4cf6=*; __utma=30149280.232591547.1629936489.1629936489.1629936489.1; __utmb=30149280.0.10.1629936489; __utmc=30149280; __utmz=30149280.1629936489.1.1.utmcsr=open.weixin.qq.com|utmccn=(referral)|utmcmd=referral|utmcct=/; __utma=223695111.1326858547.1629936489.1629936489.1629936489.1; __utmb=223695111.0.10.1629936489; __utmc=223695111; __utmz=223695111.1629936489.1.1.utmcsr=open.weixin.qq.com|utmccn=(referral)|utmcmd=referral|utmcct=/; push_noty_num=0; push_doumail_num=0; _vwo_uuid_v2=D8A1AD86A68FA87DAC0208F74852EA234|8c77aabd1c4b456acb80f56ce1913302; _pk_id.100001.4cf6=18e2fa2e664ebce9.1629936489.1.1629936615.1629936489.',
            f'Referer': f'https://movie.douban.com/subject/35427471/comments?start=20&limit=20&status=P&sort=new_score',
            'User-Agent': str(UserAgent().random)
        }

        resp = requests.get(url, headers=headers)
        if resp.status_code == requests.codes.ok:
            text = resp.json()['html']
            # 转html格式
            html_data = etree.HTML(text)

            # 获取当前页面所有的div列表
            div_list = html_data.xpath("//div[@class='comment-item ']")

            for div in div_list:
                # 姓名
                cmt_name = div.xpath("./div[@class='comment']/h3/span[@class='comment-info']/a/text()")
                cmt_names = ''.join(cmt_name).strip()

                # 评论时间
                cmt_time = div.xpath("./div[@class='comment']/h3/span[@class='comment-info']/span[@class='comment-time ']/text()")
                cmt_times = ''.join(cmt_time).strip()

                # 星评
                star = div.xpath("./div[@class='comment']/h3/span[@class='comment-info']/span[2]/@title")
                stars = ''.join(star).strip()

                # 评分
                score = div.xpath("./div[@class='comment']/h3/span[@class='comment-info']/span[2]/@class")
                scores = ''.join(score).strip(' ')[7:8]

                # 赞同人数
                vote = div.xpath("./div[@class='comment']/h3/span[@class='comment-vote']/span[@class='votes vote-count']/text()")
                votes = ''.join(vote).strip()

                # 评论内容
                comment = div.xpath("./div[@class='comment']/p[@class=' comment-content']/span[@class='short']/text()")
                comments = ''.join(comment).strip()

                wb.cell(row=count, column=1, value=cmt_names)
                wb.cell(row=count, column=2, value=cmt_times)
                wb.cell(row=count, column=3, value=stars)
                wb.cell(row=count, column=4, value=scores)
                wb.cell(row=count, column=5, value=votes)
                wb.cell(row=count, column=6, value=comments)

                count += 1

                print(cmt_names, cmt_times, stars, votes, scores, comments)
    ws.save('扫黑风暴.xlsx')


# 词云图展示
def ciyun_visual():

    rcv_data = pd.read_excel('./扫黑风暴.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    c_title = exist_col['评论'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-child',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=['的', '和', '不', '了', '是', '还', '就', '在', '很', '太', '他', '又', '能', '吧', '没', '我', '也', '有', '都', '把', '说', '人', '被', '到', '啊', '还是', '不是', '有点', '这', '你', '多', '跟', '但', '']
                   )
    print('绘图成功！')


def score_visual():
    # 读取数据
    pd_data = pd.read_excel('./扫黑风暴.xlsx')

    score = pd_data['评分'].value_counts()

    # 型号分类
    score1 = score.index.tolist()
    # 分类数据统计
    score2 = score.tolist()
    print(score1)
    print(score2)

    return score1, score2


def data_visual():
    x_data, y_data = score_visual()
    x_data = ['四星', '五星', '三星', '二星', '一星', '尚未打分']
    c = (
        Bar()
            .add_xaxis(x_data)
            .add_yaxis("评分展示", y_data, category_gap="60%")
            .set_series_opts(
            itemstyle_opts={
                "normal": {
                    "color": JsCode(
                        """new echarts.graphic.LinearGradient(0, 0, 0, 1, [{
                    offset: 0,
                    color: 'rgba(0, 244, 255, 1)'
                }, {
                    offset: 1,
                    color: 'rgba(0, 77, 167, 1)'
                }], false)"""
                    ),
                    "barBorderRadius": [30, 30, 30, 30],
                    "shadowColor": "rgb(0, 160, 221)",
                }
            }
        )
            .set_global_opts(title_opts=opts.TitleOpts(title="Bar-评分展示"))
            .render("bar_border_radius.html")
    )
    print('绘图完成！')

'''
    x_data = ['四星', '五星', '三星', '二星', '一星', '尚未打分']
    inner_data_pair = [list(z) for z in zip(x_data, y_data)]

    outer_x_data = ['四星', '五星', '三星', '二星', '一星', '尚未打分']
    y_data = [335, 310, 234, 135, 1048, 251, 147, 102]
    outer_data_pair = [list(z) for z in zip(outer_x_data, y_data)]

    (
        Pie(init_opts=opts.InitOpts(width="1600px", height="800px"))
            .add(
            series_name="访问来源",
            data_pair=inner_data_pair,
            radius=[0, "30%"],
            label_opts=opts.LabelOpts(position="inner"),
        )
            .add(
            series_name="访问来源",
            radius=["40%", "55%"],
            data_pair=outer_data_pair,
            label_opts=opts.LabelOpts(
                position="outside",
                formatter="{a|{a}}{abg|}\n{hr|}\n {b|{b}: }{c}  {per|{d}%}  ",
                background_color="#eee",
                border_color="#aaa",
                border_width=1,
                border_radius=4,
                rich={
                    "a": {"color": "#999", "lineHeight": 22, "align": "center"},
                    "abg": {
                        "backgroundColor": "#e3e3e3",
                        "width": "100%",
                        "align": "right",
                        "height": 22,
                        "borderRadius": [4, 4, 0, 0],
                    },
                    "hr": {
                        "borderColor": "#aaa",
                        "width": "100%",
                        "borderWidth": 0.5,
                        "height": 0,
                    },
                    "b": {"fontSize": 16, "lineHeight": 33},
                    "per": {
                        "color": "#eee",
                        "backgroundColor": "#334455",
                        "padding": [2, 4],
                        "borderRadius": 2,
                    },
                },
            ),
        )
            .set_global_opts(legend_opts=opts.LegendOpts(pos_left="left", orient="vertical"))
            .set_series_opts(
            tooltip_opts=opts.TooltipOpts(
                trigger="item", formatter="{a} <br/>{b}: {c} ({d}%)"
            )
        )
            .render("评分展示.html")
    )
'''

def max_data():
    # 读取数据
    pd_data = pd.read_excel('./扫黑风暴.xlsx')
    pd.set_option('display.max_columns', None)  # 显示完整的列
    pd.set_option('display.max_rows', None)  # 显示完整的行
    pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据

    # 最多播放
    max_score = pd_data[pd_data['赞同'] == pd_data['赞同'].max()]
    print(max_score)

if __name__ == '__main__':
    #spider_page()
    #ciyun_visual()
    #data_visual()
    #max_data()
    score_visual()