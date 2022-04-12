import requests
import re
import openpyxl as op
import pandas as pd
from stylecloud import gen_stylecloud
import jieba
from collections import Counter
from icecream import ic


def spider_wb():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='用户id')
    wb.cell(row=1, column=2, value='作者名称')
    wb.cell(row=1, column=3, value='作者座右铭')
    wb.cell(row=1, column=4, value='发帖时间')
    wb.cell(row=1, column=5, value='发帖内容')

    count = 2

    # 爬取第一页
    print(f'----------正在打印第1页数据----------')
    #url = 'https://m.weibo.cn/comments/hotflow?id=4661454088968882&mid=4661454088968882&max_id_type=0'
    url = 'https://m.weibo.cn/comments/hotflow?id=4690050909209999&mid=4690050909209999&max_id_type=0'
    print('当前url是：', url)

    headers = {
        'cookie': 'WEIBOCN_FROM=1110006030; SUB=_2A25MZ_SXDeRhGeVG7lAZ9S_PwjiIHXVvq5zfrDV6PUJbkdCOLUOtkW1NT7e8qp27GOSnSoETtSb_elCC-bWgVj4i; MLOGIN=1; _T_WM=38762224721; XSRF-TOKEN=d85280; M_WEIBOCN_PARAMS=oid%3D4690050909209999%26luicode%3D10000011%26lfid%3D231522type%253D1%2526t%253D10%2526q%253D%2523%25E8%258B%25B1%25E9%259B%2584%25E8%2581%2594%25E7%259B%259F%25E6%2589%258B%25E6%25B8%25B8%25E8%2583%25BD%25E7%25A2%25BE%25E5%258E%258B%25E7%258E%258B%25E8%2580%2585%25E8%258D%25A3%25E8%2580%2580%25E5%2590%2597%2523%26uicode%3D20000061%26fid%3D4690050909209999',
        'referer': 'https://m.weibo.cn/detail/4690050909209999',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4651.0 Safari/537.36'
    }

    resp = requests.get(url, headers=headers).json()
    wb_info = resp['data']['data']

    for item in wb_info:
        user_id = item.get('user')['id']  # 用户id

        author = item['user']['screen_name']  # 作者名称

        auth_sign = item['user']['description']  # 作者座右铭

        time = str(item['created_at']).split(' ')[1:4]
        rls_time = '-'.join(time)  # 发帖时间

        text = ''.join(re.findall('[\u4e00-\u9fa5]', item['text']))  # 发帖内容

        wb.cell(row=count, column=1, value=user_id)
        wb.cell(row=count, column=2, value=author)
        wb.cell(row=count, column=3, value=auth_sign)
        wb.cell(row=count, column=4, value=rls_time)
        wb.cell(row=count, column=5, value=text)
        print(user_id, author, auth_sign, rls_time, text)
        count += 1

    # 爬取第一页往后
    max_id = resp['data']['max_id']
    count = 17
    for page in range(2, 50 + 1):
        print(f'----------正在打印第{page}页数据----------')
        if page > 16:
            max_id_type = 1
        else:
            max_id_type = 0
        #url = f'https://m.weibo.cn/comments/hotflow?id=4661454088968882&mid=4661454088968882&max_id={max_id}&max_id_type={max_id_type}'
        url = f'https://m.weibo.cn/comments/hotflow?id=4690050909209999&mid=4690050909209999&max_id={max_id}&max_id_type={max_id_type}'
        print('当前url是：', url)

        resp = requests.get(url, headers=headers)

        if resp.status_code == 200:
            comments = resp.json()

            if comments['ok'] == 1:
                max_id = comments['data']['max_id']
                wb_info = comments['data']['data']

                for item in wb_info:
                    user_id = item.get('user')['id']  # 用户id

                    author = item['user']['screen_name']  # 作者名称

                    auth_sign = item['user']['description']  # 作者座右铭

                    time = str(item['created_at']).split(' ')[1:4]
                    rls_time = '-'.join(time)  # 发帖时间

                    text = ''.join(re.findall('[\u4e00-\u9fa5]', item['text']))  # 发帖内容


                    wb.cell(row=count, column=1, value=user_id)
                    wb.cell(row=count, column=2, value=author)
                    wb.cell(row=count, column=3, value=auth_sign)
                    wb.cell(row=count, column=4, value=rls_time)
                    wb.cell(row=count, column=5, value=text)
                    count += 1
                    print(user_id, author, auth_sign, rls_time, text)

                page += 1
    ws.save('英雄联盟.xlsx')


def data_visual():

    rcv_data = pd.read_excel('./鸿星尔克.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    c_title = exist_col['发帖内容'].tolist()

    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ''.join(wordlist)
    print(result)

    # 设置停用词
    stopwords = ['你', '我', '的', '了', '在', '吧', '相信', '是', '也', '都', '不', '吗', '就', '我们', '还', '大家', '你们', '就是', '以后']

    # 绘制词云图
    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-comment-dots',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=stopwords
                   )
    print('绘图成功！')



if __name__ == '__main__':
    spider_wb()
    #data_visual()