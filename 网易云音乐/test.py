from selenium import webdriver
import time
import re
from selenium.webdriver.common.by import By
import numpy as np
import pandas as pd
import random
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import os
from openpyxl.utils import get_column_letter
import requests

url = 'https://music.163.com/#/song?id=526412160'
def get_browser(url):
    chrome_options = webdriver.ChromeOptions()

    # 从 Chrome 59 版本, 支持 Headless 模式(无界面模式), 即不会弹出浏览器
    chrome_options.add_argument('--headless')

    browser = webdriver.Chrome(executable_path='./chromedriver.exe', options=chrome_options)

    browser.get(url)

    # 等待 5 秒, 让评论数据加载完成
    time.sleep(5)

    # 页面嵌套一层 iframe, 必须切换到 iframe, 才能定位的到 iframe 里面的元素
    iframe = browser.find_element_by_id('g_iframe')
    # iframe = brower.find_element_by_class_name('g-iframe')
    browser.switch_to.frame(iframe)

    # 获取【最新评论】总数
    # 参数为查找方式和值
    new_comments = browser.find_elements(By.XPATH, "//h3[@class='u-hd4']")[1].text
    # 或者
    # new_comments = brower.find_elements(By.XPATH, "//h3[@class='u-hd4'][2]")[0].text
    return browser, new_comments

browser, new_comments = get_browser(url)


def get_max_page(browser, new_comments):
    # 每页显示20条最新评论
    print('获取分页个数')
    offset = 20
    max_page = np.ceil(int(re.search('(\d{1,})', new_comments).group()) / offset)
    print('一共有', int(max_page), '个分页')
    return max_page


max_page = get_max_page(browser, new_comments)

def save_img(url, path='./1.jpg'):
    response = requests.get(url, stream=True)
    with open(path, 'wb') as f:
        f.write(response.content)

xlpath = './music.xlsx'
columns = ['head_profit', 'user_id', 'user_name', 'content', 'num_thump_up']


def insert_into_excel(data, xlpath=xlpath, columns=columns):
    """文字图片写入excel"""
    if not os.path.exists(xlpath):
        # 新建excel文件
        wb = Workbook()
        # ws1 = wb.create_sheet('music')
        # grab the active worksheet
        ws = wb.active
        ws.title = 'music'
        ws.append(columns)
    else:
        # 打开excel
        wb = load_workbook(xlpath)
        ws = wb['music']

    head_address = data[0]
    data[0] = ' '
    save_img(url=head_address, path='./1.jpg')
    ws.append(data)

    # 调整列宽
    ws.column_dimensions[get_column_letter(1)].width = 50
    # 调整行高
    ws.row_dimensions[ws.max_row].height = 50

    img = Image(r'./1.jpg')
    img.width, img.height = (50, 50)
    # ws.cell(row=3,column=1).value = " "
    img_location = f'{get_column_letter(1)}{ws.max_row}'
    ws[img_location] = " "
    ws.add_image(img, img_location)
    wb.save(xlpath)
    wb.close()


def get_comments(browser, is_first=False):
    items = browser.find_elements(By.XPATH, "//div[@class='cmmts j-flag']/div[@class='itm']")
    # 首页的数据中包含 15 条精彩评论, 20 条最新评论, 只保留最新评论
    if is_first:
        items = items[15:]

    for item in items:
        # 用户id
        user_id = item.find_elements(By.XPATH, "./div[@class='head']/a")[0].get_attribute('href').split('=')[1]
        # 用户昵称
        user_name = item.find_elements(By.XPATH, "./div[@class='cntwrap']/div[1]/div[1]/a")[0].text
        # 评论内容并去除中文冒号
        content = item.find_elements(By.XPATH, "./div[@class='cntwrap']/div[1]/div[1]")[0].text.split('：')[1]
        # 点赞数
        num_thump_up = item.find_elements(By.XPATH, "./div[@class='cntwrap']/div[@class='rp']/a[1]")[0].text

        if num_thump_up:
            num_thump_up = int(re.search('(\d{1,})', num_thump_up).group())
        else:
            num_thump_up = 0

        # 头像地址
        head_address = item.find_elements(By.XPATH, "./div[@class='head']/a/img")[-1].get_attribute('src')
        data = [head_address, user_id, user_name, content, num_thump_up]

        # 插入excel
        insert_into_excel(data, xlpath=xlpath, columns=columns)

def go_next_page(browser):
    # 模拟人为操作, 点击【下一页】(找到按钮并点击)
    next_button = browser.find_elements(By.XPATH, "//div[@class='m-cmmt']//a")[-1]
    if next_button.text == '下一页':
        next_button.click()

current = 0
def get_all_comments(current):
    while current <= max_page:
        print('正在爬取第', current, '页的数据')
        if current == 1:
            is_first = True
        else:
            is_first = False
        get_comments(browser, is_first)
        time.sleep(5)
        go_next_page(browser)
        # 模拟人为浏览
        time.sleep(random.randint(8, 12))
        current += 1
        # 记录页数
        with open('./1.txt', 'w+', encoding='utf8') as f:
            f.write(str(current))

def restart_program(browser, current):
    pages = browser.find_elements(By.XPATH, "//div[@class='m-cmmt']/div[3]/div/a")
    # previous_button = pages[0]
    # next_button = pages[-1]
    page_num = []
    for page in pages[1:-1]:
        # 查找页数
        try:
            page_num.append(int(page.text))
        except:
            # pages中的'...'用-1表示
            page_num.append(-1)
    if current in page_num:
        index = page_num.index(current) + 1
        pages[index].click()
        get_all_comments(current)
    else:
        page_second_max = np.array(page_num).argsort()[-2]
        index = page_second_max + 1
        pages[index].click()
        restart_program(browser, current)


if __name__ == '__main__':
    url = 'https://music.163.com/#/song?id=526412160'
    browser, new_comments = get_browser(url)
    get_max_page(browser, new_comments)

    get_comments(browser)
    go_next_page(browser)
    restart_program(browser, get_all_comments(0))
