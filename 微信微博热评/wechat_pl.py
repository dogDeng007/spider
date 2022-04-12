'''
https://m.weibo.cn/comments/hotflow?id=4666743534650015&mid=4666743534650015&max_id_type=0

'''


import requests
import re
import openpyxl as op
import jieba
from icecream import ic
from stylecloud import gen_stylecloud
from collections import Counter
from pyecharts import options as opts
from pyecharts.charts import Bar
import pandas as pd
from snownlp import SnowNLP


def spider_wb():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='用户id')
    wb.cell(row=1, column=2, value='作者名称')
    wb.cell(row=1, column=3, value='作者座右铭')
    wb.cell(row=1, column=4, value='发帖时间')
    wb.cell(row=1, column=5, value='点赞人数')
    wb.cell(row=1, column=6, value='发帖内容')

    count = 2

    # 爬取第一页
    print(f'----------正在打印第1页数据----------')
    url = 'https://m.weibo.cn/comments/hotflow?id=4677708824712370&mid=4677708824712370&max_id_type=0'
    print('当前url是：', url)

    headers = {
        'cookie': 'SUB=_2A25NyTOqDeRhGeVG7lAZ9S_PwjiIHXVvMl3irDV6PUJbktB-LVDmkW1NT7e8qozwK1pqWVKX_PsKk5dhdCyPXwW1; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WFGibRIp_iSfMUfmcr5kb295NHD95Q01h-E1h-pe0.XWs4DqcjLi--fi-2Xi-2Ni--fi-z7iKysi--Ri-8si-zXi--fi-88i-zce7tt; _T_WM=98961943286; MLOGIN=1; WEIBOCN_FROM=1110006030; XSRF-TOKEN=70a1e0; M_WEIBOCN_PARAMS=oid%3D4648381753067388%26luicode%3D20000061%26lfid%3D4648381753067388%26uicode%3D20000061%26fid%3D4648381753067388',
        'Referer': 'https://m.weibo.cn/detail/4666743534650015',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4573.0 Safari/537.36'
    }

    resp = requests.get(url, headers=headers).json()

    wb_info = resp['data']['data']

    for item in wb_info:

        # 用户id
        user_id = item.get('user')['id']

        # 作者名称
        author = item['user']['screen_name']

        #作者座右铭
        auth_sign = item['user']['description']

        # 发帖时间
        time = str(item['created_at']).split(' ')[1:4]
        rls_time = '-'.join(time)

        # 点赞人数
        stars = item.get('like_count')

        # 发帖内容
        text = ''.join(re.findall('[\u4e00-\u9fa5]', item['text']))

        print(user_id, author, auth_sign, rls_time, stars, text)

        wb.cell(row=count, column=1, value=user_id)
        wb.cell(row=count, column=2, value=author)
        wb.cell(row=count, column=3, value=auth_sign)
        wb.cell(row=count, column=4, value=rls_time)
        wb.cell(row=count, column=5, value=stars)
        wb.cell(row=count, column=6, value=text)
        count += 1

    # 爬取第一页往后
    max_id = resp['data']['max_id']
    count = 17
    for page in range(2, 100 + 1):
        print(f'----------正在打印第{page}页数据----------')
        if page > 16:
            max_id_type = 1
        else:
            max_id_type = 0
        url = f'https://m.weibo.cn/comments/hotflow?id=4677708824712370&mid=4677708824712370&max_id={max_id}&max_id_type={max_id_type}'
        print('当前url是：', url)

        resp = requests.get(url, headers=headers)

        if resp.status_code == 200:
            comments = resp.json()

            if comments['ok'] == 1:
                max_id = comments['data']['max_id']
                wb_info = comments['data']['data']

                for item in wb_info:

                    # 用户id
                    user_id = item.get('user')['id']

                    # 作者名称
                    author = item['user']['screen_name']

                    # 作者座右铭
                    auth_sign = item['user']['description']

                    # 发帖时间
                    time = str(item['created_at']).split(' ')[1:4]

                    rls_time = '-'.join(time)

                    # 点赞人数
                    stars = item.get('like_count')

                    text = ''.join(re.findall('[\u4e00-\u9fa5]', item['text']))  # 发帖内容


                    wb.cell(row=count, column=1, value=user_id)
                    wb.cell(row=count, column=2, value=author)
                    wb.cell(row=count, column=3, value=auth_sign)
                    wb.cell(row=count, column=4, value=rls_time)
                    wb.cell(row=count, column=5, value=stars)
                    wb.cell(row=count, column=6, value=text)
                    count += 1
                    print(user_id, author, auth_sign, rls_time, stars, text)

                page += 1
    ws.save('微信收费.xlsx')
    print('数据保存完毕！')


def visual_ciyun():
    # 读取文件
    rcv_data = pd.read_excel('./微信收费.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    c_title = exist_col['发帖内容'].tolist()

    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)

    # 设置停用词
    stop_words = ['的', '了', '我', '也', '要', '吧', '吗', '那', '用', '就', '是', '想', '在', '有', '哪', '啊', '这', '你', '还', '又', '见', '给', '可以', '微笑']

    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fab fa-google',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=stop_words
                   )
    print('词云图绘制成功！')

    # 词频设置
    all_words = [word for word in result.split(' ') if len(word) > 1 and word not in stop_words]
    wordcount = Counter(all_words).most_common(10)

    x_data, y_data = list(zip(*wordcount))
    print(x_data)
    print(y_data)

    c = (
        Bar()
            .add_xaxis(x_data)
            .add_yaxis("微信收费-词频展示图", y_data)
            .set_global_opts(title_opts=opts.TitleOpts(title="微信-词频展示图"))
            .set_series_opts(
            label_opts=opts.LabelOpts(is_show=False),
            markpoint_opts=opts.MarkPointOpts(
                data=[
                    opts.MarkPointItem(type_="max", name="最大值"),
                    opts.MarkPointItem(type_="min", name="最小值"),
                    opts.MarkPointItem(type_="average", name="平均值"),
                ]
            ),
        )
            .render("微信词频展示.html")
    )

    c = (
        Bar()
            .add_xaxis(x_data)
            .add_yaxis("微信-词频展示图", y_data)
            .set_global_opts(
            title_opts=opts.TitleOpts(title="微信-词频展示图"),
            toolbox_opts=opts.ToolboxOpts(),
            legend_opts=opts.LegendOpts(is_show=False),
        )
            .render("微信-词频展示图.html")
    )
    print('微信-词频展示图绘制成功！')


def max_data():
    # 读取数据
    pd_data = pd.read_excel('./微信收费.xlsx')
    pd.set_option('display.max_columns', None)  # 显示完整的列
    pd.set_option('display.max_rows', None)  # 显示完整的行
    pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据

    # 最多点赞数
    max_star = pd_data[pd_data['点赞人数'] == pd_data['点赞人数'].max()]
    print(max_star)

    # 情感分析
    pinglun = '如果这样那都去支付宝吧反正功能一样的'

    Sentiment_analysis = SnowNLP(pinglun).sentiments
    print(Sentiment_analysis)


if __name__ == '__main__':
    spider_wb()
    #visual_ciyun()
    #max_data()