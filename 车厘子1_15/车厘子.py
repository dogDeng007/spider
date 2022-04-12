'''
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=20180186520&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=20180186520&score=0&sortType=5&page=1&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=20180186520&score=0&sortType=5&page=2&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=20180186520&score=0&sortType=5&page=3&pageSize=10&isShadowSku=0&rid=0&fold=1
'''

import requests
import json
import openpyxl as op
import pandas as pd
import jieba
from stylecloud import gen_stylecloud
from collections import Counter
from snownlp import SnowNLP
from pyecharts.charts import Bar
import pyecharts.options as opts
from pyecharts.charts import Line, Grid
from icecream import ic

def spider_jd():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='昵称')
    wb.cell(row=1, column=2, value='评分')
    wb.cell(row=1, column=3, value='产品类型')
    wb.cell(row=1, column=4, value='评论时间')
    wb.cell(row=1, column=5, value='评论点赞数')
    wb.cell(row=1, column=6, value='评论回复数')
    wb.cell(row=1, column=7, value='评论内容')
    count = 2

    for page in range(1, 70+1):

        print(f'-----------------正在抓取第{page}页数据中-----------------')
        url = f'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=20180186520&score=0&sortType=5&page={page}&pageSize=10&isShadowSku=0&rid=0&fold=1'

        headers = {
            'Cookie': 'unpl=JF8EAK1nNSttUEhVABMFSREXSAkAWwpbGUcAO2RQAF9dQgZWSQdMQBJ7XlVdXhRKFx9vZBRUXlNJVQ4YCisSEXteXVdZDEsWC2tXVgQFDQ8VXURJQlZAFDNVCV9dSRZRZjJWBFtdT1xWSAYYRRMfDlAKDlhCR1FpMjVkXlh7VAQrARgVEkteVFdeOHsQM19XDF1ZXElcNRoyGiJSHwFSWV4PThFObGQCVl1bS10GKwMrEQ; __jdu=226312182; shshshfpa=efbab15b-4e59-4172-3853-0814c482e5b7-1641961835; shshshfpb=zE_mfsMvPvmTeYovmWU8U6A; __jdv=76161171|baidu|-|organic|notset|1642207402772; shshshfp=6a2f9d430c5472d0e4d6987a07b5e650; __jda=122270672.226312182.1641961830.1641961832.1642207403.2; __jdc=122270672; token=520eeef0e875bc8fa6f12b0e79242479,2,912337; __tk=uSGEYzoovUkEXpdSYUGFuDawXSvTXpG0vpvTX3uFXSX,2,912337; ip_cityCode=2376; areaId=1; ipLoc-djd=1-72-55653-0; shshshsID=e16908b9e1521c6376977cddb6673235_4_1642207447178; __jdb=122270672.4.226312182|2.1642207403; 3AB9D23F7A4B3C9B=DJJ566ILWBSVQYWTANAILV7253S3BISPU4VGZSPRN6BLKTNZMV6CAHRONFI5O3HWVKHQBVF2LI7FM2DWFMJ7JBTW7U; jwotest_product=99; JSESSIONID=2AACBB960318D34406A1F826D05A0382.s1',
            'Referer': 'https://item.jd.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'
        }

        # 获取响应
        resp = requests.get(url, headers=headers)
        ic(resp.text)
        json_data = json.loads(resp.text[20:-2])


        for cmts in json_data['comments']:

            # 昵称
            nickname = cmts['nickname']
            # 评分
            score = cmts['score']
            # 评论
            comments = cmts['content']
            # 产品类型
            product = cmts['productColor']

            # 评论时间
            time = cmts['referenceTime']

            # 评论点赞数
            starVote = cmts['usefulVoteCount']

            # 评论回复数
            cmtsReply = cmts['replyCount']

            wb.cell(row=count, column=1, value=nickname)
            wb.cell(row=count, column=2, value=score)
            wb.cell(row=count, column=3, value=product)
            wb.cell(row=count, column=4, value=time)
            wb.cell(row=count, column=5, value=starVote)
            wb.cell(row=count, column=6, value=cmtsReply)
            wb.cell(row=count, column=7, value=comments)


            print(nickname, score, comments, product, time, starVote, cmtsReply)
            count += 1

    ws.save('车厘子.xlsx')


if __name__ == '__main__':
    spider_jd()
    #visual_ciyun()
    #product_visual()
    #max_data()
    #extr_time()