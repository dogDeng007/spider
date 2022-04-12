'''
https://www.ixigua.com/tlb/comment/article/v5/tab_comments/?tab_index=0&count=10&group_id=6991462962133369380&item_id=6991462962133369380&aid=1768
https://www.ixigua.com/tlb/comment/article/v5/tab_comments/?tab_index=0&count=10&offset=10&group_id=6991462962133369380&item_id=6991462962133369380&aid=1768
https://www.ixigua.com/tlb/comment/article/v5/tab_comments/?tab_index=0&count=10&offset=20&group_id=6991462962133369380&item_id=6991462962133369380&aid=1768
https://www.ixigua.com/tlb/comment/article/v5/tab_comments/?tab_index=0&count=10&offset=30&group_id=6991462962133369380&item_id=6991462962133369380&aid=1768
https://www.ixigua.com/tlb/comment/article/v5/tab_comments/?tab_index=0&count=10&offset=40&group_id=6991462962133369380&item_id=6991462962133369380&aid=1768
'''

import requests
import time
import openpyxl as op
from icecream import ic
import pandas as pd
import jieba
def spider_page():
    # 创建workbook
    ws = op.Workbook()
    # 创建worksheet
    wb = ws.create_sheet(index=0)

    # 创建表头
    wb.cell(row=1, column=1, value='用户名称')
    wb.cell(row=1, column=2, value='评论点赞')
    wb.cell(row=1, column=3, value='评论时间')
    wb.cell(row=1, column=4, value='贴子回复')
    wb.cell(row=1, column=5, value='评论内容')
    count = 2

    for page in range(1, 200+1):
        print(f'--------------------------------正在打印第{page}页--------------------------------')
        url = f'https://www.ixigua.com/tlb/comment/article/v5/tab_comments/?tab_index=0&count=10&offset={(page-1)*10}&group_id=6991462962133369380&item_id=6991462962133369380&aid=1768'
        headers = {
            'cookie': 'MONITOR_WEB_ID=08d6862a-9750-4278-adde-e7a504168f20; __ac_nonce=0610a443f007305b4e54d; __ac_signature=_02B4Z6wo00f01eNI4BQAAIDAgEIgfmddvGHjaOSAABnaa9; ttcid=036bb950feac4656b5c528086180cd6727; passport_csrf_token_default=a79c96dc2fe9736eb5dde67994394098; sid_guard=35fab0e1a6c63353f66d9a21f5988e87%7C1628062786%7C3023297%7CWed%2C+08-Sep-2021+07%3A28%3A03+GMT; uid_tt=2bfe13162d1e63ecb4fcf1b616bc7f41; uid_tt_ss=2bfe13162d1e63ecb4fcf1b616bc7f41; sid_tt=35fab0e1a6c63353f66d9a21f5988e87; sessionid=35fab0e1a6c63353f66d9a21f5988e87; sessionid_ss=35fab0e1a6c63353f66d9a21f5988e87; sid_ucp_v1=1.0.0-KGZmZjk2NDNjZjk2ZDUyZmI5YjM3MWQ4YjRiOWZiNWY5NGRhYmRkZmQKEAjR3-qG-QIQwoipiAYY6A0aAmxmIiAzNWZhYjBlMWE2YzYzMzUzZjY2ZDlhMjFmNTk4OGU4Nw; ssid_ucp_v1=1.0.0-KGZmZjk2NDNjZjk2ZDUyZmI5YjM3MWQ4YjRiOWZiNWY5NGRhYmRkZmQKEAjR3-qG-QIQwoipiAYY6A0aAmxmIiAzNWZhYjBlMWE2YzYzMzUzZjY2ZDlhMjFmNTk4OGU4Nw; passport_csrf_token=a79c96dc2fe9736eb5dde67994394098; ixigua-a-s=1; ttwid=1%7CQ4r5hI0DCyQV6RHSXmCmda5wfP0AXHtbeYfj1GFzPMI%7C1628063087%7Cfc880dbc2e6fe424f49ebecdb1182fdb3358372d95bb22db1a83e47a11d23657',
            'referer': 'https://www.ixigua.com/6991462962133369380?is_new_connect=0&is_new_user=0',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.8 Safari/537.36'
        }

        resp = requests.get(url, headers = headers)

        if resp.status_code == requests.codes.ok:

            json_data = resp.json()['data']

            for item in json_data:

                # 用户名称
                user = item['comment']['user_name']

                # 评论内容
                text = item['comment']['text']

                # 贴子回复数
                reply = item['comment']['reply_count']

                # 评论时间
                times = item['comment']['create_time']
                print(times)
                print(type(times))
                rls_time = time.strftime('%Y-%m-%d %H:%M', time.localtime(times))

                # 评论点赞数
                stars = item['comment']['digg_count']

                wb.cell(row=count, column=1, value=user)
                wb.cell(row=count, column=2, value=stars)
                wb.cell(row=count, column=3, value=rls_time)
                wb.cell(row=count, column=4, value=reply)
                wb.cell(row=count, column=5, value=text)

                count += 1
                ic(user, stars, rls_time, reply, text)
    # 保存数据
    ws.save('今日头条1.xlsx')
    print('数据保存完毕！')


# 读取源文件
rcv_data = pd.read_excel('今日头条1.xlsx')

# 词云图展示
def view_ciyun():


    # 删除空行
    exist_col = rcv_data.dropna()
    c_title = exist_col['评论内容'].tolist()

    # jieba分词
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img.jpg'

    # 生成词云图
    gen_stylecloud(text=result,
                   icon_name='fas fa-hourglass',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=['的', '是', '了', '你', '才', '有', '我', '这', '他', '在', '人', '啊', '都', '说', '太',
                                      '谁', '什么', '不', '一个', '没', '就',  '捂脸', '看', '吗', '要', '很', '应该', '这么', '这个']
                   )
    print('绘图成功！')

def datas_anay():
    max_stars = rcv_data[rcv_data['评论点赞'] == rcv_data['评论点赞'].max()]
    ic(max_stars)

    max_reply = rcv_data[rcv_data['贴子回复'] == rcv_data['贴子回复'].max()]
    ic(max_reply)


if __name__ == '__main__':
    spider_page()
    #view_ciyun()
    #datas_anay()




