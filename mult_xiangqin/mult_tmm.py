import requests
from queue import Queue
import threading
from fake_useragent import UserAgent
from lxml import etree
import openpyxl as op
import os
import shutil
import requests
from fake_useragent import UserAgent
import time, random
from faker import Faker
from pyecharts.charts import Funnel, Page  # 各个图形的类
from pyecharts.charts import Line
import pandas as pd
import jieba
from stylecloud import gen_stylecloud
from pyecharts.charts import Bar,Pie
from pyecharts import options as opts


from pyecharts.charts import Bar, Pie, Funnel, Page  # 各个图形的类
from pyecharts.charts import Line
import pandas as pd
from pyecharts import options as opts


pic_path = './pictures'
class Procuder(threading.Thread):
    headers = {
        'cookie': 'mediav=%7B%22eid%22%3A%22810564%22%2C%22ep%22%3A%22%22%2C%22vid%22%3A%225pxUu%3El9j%258odUSi%3CsF9%22%2C%22ctn%22%3A%22%22%2C%22vvid%22%3A%225pxUu%3El9j%258odUSi%3CsF9%22%2C%22_mvnf%22%3A1%2C%22_mvctn%22%3A0%2C%22_mvck%22%3A0%2C%22_refnf%22%3A0%7D; Qs_lvt_331562=1625187114%2C1625216478%2C1625220965; user_sid=c353fd46d8; Hm_lvt_484b42ae470fa380c479571505f92406=1625187114,1625220966,1625229265; mediav=%7B%22eid%22%3A%22810564%22%2C%22ep%22%3A%22%22%2C%22vid%22%3A%225pxUu%3El9j%258odUSi%3CsF9%22%2C%22ctn%22%3A%22%22%2C%22vvid%22%3A%225pxUu%3El9j%258odUSi%3CsF9%22%2C%22_mvnf%22%3A1%2C%22_mvctn%22%3A0%2C%22_mvck%22%3A0%2C%22_refnf%22%3A1%7D; nb-referrer-hostname=www.csflhjw.com; Qs_pv_331562=2666523269737035300%2C1447382052269964000%2C787150967495128200%2C927396808731292400%2C733849117706530800; Hm_lpvt_484b42ae470fa380c479571505f92406=1625229584; nb-start-page-url=https%3A%2F%2Fwww.csflhjw.com%2Fzhenghun%2F9.html',
        'referer': 'https://www.csflhjw.com/zhenghun/9.html',
        'User-Agent': str(UserAgent().random)
    }

    def __init__(self, page_queue, img_queue, *args, **kwargs):
        super(Procuder, self).__init__(*args, **kwargs)
        self.page_queue = page_queue
        self.img_queue = img_queue


    def run(self) -> None:
        while True:
            if self.page_queue.empty():
                break
            url = self.page_queue.get()
            self.parse_page(url)


    def parse_page(self, url):
        resp = requests.get(url, headers=self.headers)

        if resp.status_code == 200:
            html_data = etree.HTML(resp.text)
            divs = html_data.xpath("//div[@class='zh-item']/div[@class='e']")

            for div in divs:
                # 获取当前页面信息
                name = div.xpath("./div[@class='e-name']/h2/text()")  # 小姐姐称呼
                name = ''.join(name)

                infos = div.xpath(".//div[@class='e-intro']/p[1]/text()")
                bir_date = ''.join(infos).split(' ')[0]  # 小姐姐出生日期
                height = ''.join(infos).split(' ')[1]  # 小姐姐身高
                educ_bgd = ''.join(infos).split(' ')[2]  # 小姐姐学历

                mary_stus = div.xpath(".//div[@class='e-intro']/p[2]/text()")  # 小姐姐婚否
                mary_stus = ''.join(mary_stus).split('：')[1]

                profe = div.xpath(".//div[@class='e-intro']/p[3]/text()")  # 小姐姐职业
                profe = ''.join(profe).split('：')[1]

                # 获取详情页面信息
                links = div.xpath("./div[@class='e-name']/a[@class='e-a']/@href")
                x = 'https://www.csflhjw.com'
                next_page = [x + i for i in links]

                for next in next_page:
                    response = requests.get(next, headers=self.headers)

                    if response.status_code == 200:
                        html = etree.HTML(response.text)

                        prof_phot = html.xpath("//div[@class='team-img']/img/@src")  # 小姐姐照片
                        x = 'https://www.csflhjw.com'
                        prof_phot = [x + i for i in prof_phot]
                        prof_phot = ''.join(prof_phot)

                        children = html.xpath("//div[@class='team-e']/p[4]/text()")  # 有无子女
                        children = ''.join(children).split('：')[1]

                        room = html.xpath("//div[@class='team-e']/p[5]/text()")  # 是否有房
                        room = ''.join(room).split('：')[1]

                        man_age = html.xpath("//div[@class='hunyin-1-2']/p[1]/span[1]/text()")  # 择偶年龄
                        man_age = ''.join(man_age).split('龄：')[1]

                        man_city = html.xpath("//div[@class='hunyin-1-2']/p[1]/span[2]/text()")  # 择偶城市
                        man_city = ''.join(man_city).split('市：')[1]

                        man_req = html.xpath("//div[@class='hunyin-1-2']/p[2]/span/text()")  # 男方要求
                        man_req = ''.join(man_req).split('求：')[1]

                        introd_myself = html.xpath("//div[@class='hunyin-1-3']/p/text()")  # 个人独白
                        introd_myself = ''.join(introd_myself).strip()

                        self.img_queue.put((name, bir_date, height, educ_bgd, mary_stus, profe, next_page, prof_phot, children, room, man_age, man_city, man_req, introd_myself))

class Consumer(threading.Thread):

    def __init__(self, page_queue, img_queue, *args, **kwargs):
        super(Consumer, self).__init__(*args, **kwargs)
        self.page_queue = page_queue
        self.img_queue = img_queue


    def run(self) -> None:
        ws = op.Workbook()
        wb = ws.create_sheet(index=0)

        wb.cell(row=1, column=1, value='称呼')
        wb.cell(row=1, column=2, value='出生日期')
        wb.cell(row=1, column=3, value='身高')
        wb.cell(row=1, column=4, value='学历')
        wb.cell(row=1, column=5, value='婚否')
        wb.cell(row=1, column=6, value='职业')
        wb.cell(row=1, column=7, value='有无子女')
        wb.cell(row=1, column=8, value='是否有房')
        wb.cell(row=1, column=9, value='择偶年龄')
        wb.cell(row=1, column=10, value='择偶城市')
        wb.cell(row=1, column=11, value='男方要求')
        wb.cell(row=1, column=12, value='个人独白')
        count = 2

        while True:
            if self.img_queue.empty() and self.page_queue.empty():
                break
            name, bir_date, height, educ_bgd, mary_stus, profe, next_page, prof_phot, children, room, man_age, man_city, man_req, introd_myself = self.img_queue.get()

            wb.cell(row=count, column=1, value=name)
            wb.cell(row=count, column=2, value=bir_date)
            wb.cell(row=count, column=3, value=height)
            wb.cell(row=count, column=4, value=educ_bgd)
            wb.cell(row=count, column=5, value=mary_stus)
            wb.cell(row=count, column=6, value=profe)
            wb.cell(row=count, column=7, value=children)
            wb.cell(row=count, column=8, value=room)
            wb.cell(row=count, column=9, value=man_age)
            wb.cell(row=count, column=10, value=man_city)
            wb.cell(row=count, column=11, value=man_req)
            wb.cell(row=count, column=12, value=introd_myself)

            print(prof_phot, name, bir_date, height, educ_bgd, mary_stus, profe, children, room,
                  man_age, man_city, man_req, introd_myself)
            count += 1
            ws.save('相亲网小姐姐666.xlsx')

def main():
    page_queue = Queue(100)
    img_queue = Queue(1000)

    for page in range(1, 20+1):
        print(f'--------------------正在打印第{page}页数据--------------------')
        url = f'https://www.csflhjw.com/zhenghun/9.html?page={page}'
        page_queue.put(url)

    # 开启5个生产者
    for x in range(5):
        t = Procuder(page_queue, img_queue)
        t.start()

    # 开启10个消费者
    for x in range(10):
        t = Consumer(page_queue, img_queue)
        t.start()
        t.join()

pd_data = pd.read_excel('相亲网小姐姐.xlsx')
print(pd_data.tail())
pd.set_option('display.max_columns', None)  # 显示完整的列
pd.set_option('display.max_rows', None)  # 显示完整的行
pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据


def anay_height():
    pd_data['身高'].dropna()  # 删除空行

    print(pd_data['身高1'])

def anay_age():
    '''
    年龄分析
    '''
    age_count = pd_data['出生日期'].value_counts()
    # 针对于出生日期的分类
    x_data = age_count.index.tolist()
    y_data = age_count.tolist()

    return x_data, y_data


# 画图
def data_visual():

    x_data, y_data = anay_height()
    print(x_data)
    print(y_data)

    # 1、柱状图
    def barPage() -> Bar:
        bar = (
            Bar()
                .add_xaxis(x_data)
                .add_yaxis("柱状图", y_data)
                .set_global_opts(
                title_opts=opts.TitleOpts(title="柱状图分析"),
                legend_opts=opts.LegendOpts(is_show=False), )
        )
        return bar

    # 2、饼图
    def piePage() -> Pie:
        pie = (
            Pie()
                .add("", [list(z) for z in zip(x_data, y_data)])
                .set_global_opts(title_opts=opts.TitleOpts(title="饼图分析"))
                .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}"))
        )
        return pie

    # 3、折线图
    def linePage() -> Line:
        line = (
            Line()
                .add_xaxis(x_data)
                .add_yaxis("折线图", y_data)
                .set_global_opts(title_opts=opts.TitleOpts(title="折线图分析"))
        )
        return line

    # 4、漏斗图
    def funnelPage() -> Funnel:
        funnel = (
            Funnel()
                .add("漏斗图", [list(z) for z in zip(x_data, y_data)])
                .set_global_opts(title_opts=opts.TitleOpts(title="漏斗图分析"))
        )
        return funnel


    # 上面是6个图形的代码，下面利用Page进行组合
    # !!! 关键步骤
    page = (
        Page(layout=Page.DraggablePageLayout)
            .add(
            barPage(),
            piePage(),
            linePage(),
            funnelPage())
    )

    page.render("小姐姐年龄展示图.shtml")
    print('绘图完成！')



def ciyun():
    pd_data = pd.read_excel('相亲网小姐姐.xlsx')

    exist_col = pd_data.dropna()        # 删除空行
    c_title = exist_col['男方要求'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-desktop',  #s
                   font_path='msyh.ttc',
                   background_color="black",
                   output_name=pic)
    print('绘图成功！')



if __name__ == '__main__':
    #anay_data()
    #ciyun()
    #data_visual()
    anay_height()