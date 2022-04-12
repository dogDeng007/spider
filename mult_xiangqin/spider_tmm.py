import time

import requests
from fake_useragent import UserAgent
from lxml import etree
import openpyxl as op
import os
import shutil


def spider_page():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='称呼')
    wb.cell(row=1, column=2, value='出生日期')
    wb.cell(row=1, column=3, value='身高')
    wb.cell(row=1, column=4, value='学历')
    wb.cell(row=1, column=5, value='婚否')
    wb.cell(row=1, column=6, value='职业')
    wb.cell(row=1, column=7, value='有无子女')
    wb.cell(row=1, column=8, value='是否有房')
    wb.cell(row=1, column=9, value='择偶年龄')
    wb.cell(row=1, column=10, value='择偶城市')
    wb.cell(row=1, column=11, value='男方要求')
    wb.cell(row=1, column=12, value='个人独白')
    count = 2

    # 爬取图片的list
    save_pic = []
    for page in range(1, 20 + 1):
        print(f'--------------------正在打印第{page}页数据--------------------')
        url = f'https://www.csflhjw.com/zhenghun/9.html?page={page}'

        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.40 Safari/537.36'
        }

        resp = requests.get(url, headers = headers)
        if resp.status_code == 200:
            html_data = etree.HTML(resp.text)
            divs = html_data.xpath("//div[@class='zh-item']/div[@class='e']")
            for div in divs:
                # 获取当前页面信息
                name = div.xpath("./div[@class='e-name']/h2/text()")  # 小姐姐称呼
                name = ''.join(name)

                infos = div.xpath(".//div[@class='e-intro']/p[1]/text()")
                bir_date = ''.join(infos).split(' ')[0]  # 小姐姐出生日期
                height = ''.join(infos).split(' ')[1]  # 小姐姐身高
                educ_bgd = ''.join(infos).split(' ')[2]  # 小姐姐学历

                mary_stus = div.xpath(".//div[@class='e-intro']/p[2]/text()")  # 小姐姐婚否
                mary_stus = ''.join(mary_stus).split('：')[1]

                profe = div.xpath(".//div[@class='e-intro']/p[3]/text()")  # 小姐姐职业
                profe = ''.join(profe).split('：')[1]


                # 获取详情页面信息
                links = div.xpath("./div[@class='e-name']/a[@class='e-a']/@href")
                x = 'https://www.csflhjw.com'
                next_page = [x + i for i in links]

                for next in next_page:
                    response = requests.get(next, headers=headers)
                    if response.status_code == 200:
                        html = etree.HTML(response.text)

                        prof_phot = html.xpath("//div[@class='team-img']/img/@src")     # 小姐姐照片
                        x = 'https://www.csflhjw.com'
                        prof_phot = [x + i for i in prof_phot]
                        prof_phot = ''.join(prof_phot)

                        children = html.xpath("//div[@class='team-e']/p[4]/text()")     # 有无子女
                        children = ''.join(children).split('：')[1]

                        room = html.xpath("//div[@class='team-e']/p[5]/text()")         # 是否有房
                        room = ''.join(room).split('：')[1]

                        man_age = html.xpath("//div[@class='hunyin-1-2']/p[1]/span[1]/text()")     # 择偶年龄
                        man_age = ''.join(man_age).split('龄：')[1]

                        man_city = html.xpath("//div[@class='hunyin-1-2']/p[1]/span[2]/text()")        # 择偶城市
                        man_city = ''.join(man_city).split('市：')[1]

                        man_req = html.xpath("//div[@class='hunyin-1-2']/p[2]/span/text()")         # 男方要求
                        man_req = ''.join(man_req).split('求：')[1]

                        introd_myself = html.xpath("//div[@class='hunyin-1-3']/p/text()")       # 个人独白
                        introd_myself = ''.join(introd_myself).strip()

                        wb.cell(row=count, column=1, value=name)
                        wb.cell(row=count, column=2, value=bir_date)
                        wb.cell(row=count, column=3, value=height)
                        wb.cell(row=count, column=4, value=educ_bgd)
                        wb.cell(row=count, column=5, value=mary_stus)
                        wb.cell(row=count, column=6, value=profe)
                        wb.cell(row=count, column=7, value=children)
                        wb.cell(row=count, column=8, value=room)
                        wb.cell(row=count, column=9, value=man_age)
                        wb.cell(row=count, column=10, value=man_city)
                        wb.cell(row=count, column=11, value=man_req)
                        wb.cell(row=count, column=12, value=introd_myself)

                        print(type(prof_phot))
                        print(type(name))
                        save_pic.append((prof_phot, name))
                        print('-----正在采集第' + str(count-1) + '条数据-----')
                        count += 1
                        print(prof_phot,name, bir_date, height, educ_bgd, mary_stus, profe, children, room,
                              man_age, man_city, man_req, introd_myself)
    ws.save('相亲网小姐姐.xlsx')
    return save_pic


# 图片保存路径
pic_path = './pictures'
def down_tmm(save_pic):
    # 判断文件夹是否存在
    if not os.path.exists(pic_path):
        os.mkdir(pic_path)
    else:
        # 判断文件夹是否为空
        shutil.rmtree(pic_path)
        os.mkdir(pic_path)

    count = 0
    for num, (prof_phot, name) in enumerate(save_pic):
        r = requests.get(prof_phot)
        pic = r.content
        try:
            with open('./pictures/{}.jpeg'.format(name), 'wb') as fin:
                print(f'正在爬取第{count}张图片')
                fin.write(pic)
                print('{}.jpeg----下载成功'.format(name))
        except:
            print('下载失败！')
        count += 1

if __name__ == '__main__':
    down_tmm(spider_page())

