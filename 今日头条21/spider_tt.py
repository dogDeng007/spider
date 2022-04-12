'''
https://www.toutiao.com/article/v2/tab_comments/?aid=24&app_name=toutiao_web&offset=0&count=20&group_id=7032951744313164295&item_id=7032951744313164295
https://www.toutiao.com/article/v2/tab_comments/?aid=24&app_name=toutiao_web&offset=20&count=20&group_id=7032951744313164295&item_id=7032951744313164295
https://www.toutiao.com/article/v2/tab_comments/?aid=24&app_name=toutiao_web&offset=40&count=20&group_id=7032951744313164295&item_id=7032951744313164295
https://www.toutiao.com/article/v2/tab_comments/?aid=24&app_name=toutiao_web&offset=60&count=20&group_id=7032951744313164295&item_id=7032951744313164295
'''

import requests
import time
import openpyxl as op
from icecream import ic
import pandas as pd
import jieba
from stylecloud import gen_stylecloud

def spider_page():
    # 创建workbook
    ws = op.Workbook()
    # 创建worksheet
    wb = ws.create_sheet(index=0)

    # 创建表头
    wb.cell(row=1, column=1, value='用户名称')
    wb.cell(row=1, column=2, value='评论点赞')
    wb.cell(row=1, column=3, value='评论时间')
    wb.cell(row=1, column=4, value='贴子回复')
    wb.cell(row=1, column=5, value='评论内容')
    count = 2

    for page in range(1, 160+1):
        print(f'--------------------------------正在打印第{page}页--------------------------------')
        url = f'https://www.toutiao.com/article/v2/tab_comments/?aid=24&app_name=toutiao_web&offset={(page-1)*20}&count=20&group_id=7032951744313164295&item_id=7032951744313164295'
        headers = {
            'cookie': 'ttcid=4491644d412e43d2b7f23c76a772570341; tt_webid=7030387321962808839; csrftoken=46e253f1c925ceb74930797c815bff92; MONITOR_DEVICE_ID=ca2ba9c9-c54c-4064-bced-c5538d400171; _S_WIN_WH=2560_1289; _S_DPR=1; _S_IPAD=0; __ac_nonce=0619a3adb00f7b69f2666; __ac_signature=_02B4Z6wo00f01s5iLxAAAIDDrWjvep6wnO7ORiuAANIP57W3fRp0nSsVzLuXOVfy70Yc4fLrlyZ5kPMf9PICnCv-MjaQ46j7H.oNopR165VhH4HQ8miqnVygq0nngm6QY.zTc3gT0oMZpdFIb0; s_v_web_id=verify_f94db67661015f89c7e044e5fd8aa500; _tea_utm_cache_2018=undefined; passport_csrf_token_default=1cef6db8c3a7ae6cab047dbb49a3097f; passport_csrf_token=1cef6db8c3a7ae6cab047dbb49a3097f; sso_uid_tt_ss=9e782a83011753f473a1597486d3a5ea; toutiao_sso_user=19a198961a7901a17e6acfe60af7f906; toutiao_sso_user_ss=19a198961a7901a17e6acfe60af7f906; sso_uid_tt=9e782a83011753f473a1597486d3a5ea; uid_tt=3a79f16ec597e54bdb9284e5c090f6b4; uid_tt_ss=3a79f16ec597e54bdb9284e5c090f6b4; sid_tt=3b25752d4503c200a1036e56aa0eecc7; sessionid=3b25752d4503c200a1036e56aa0eecc7; sessionid_ss=3b25752d4503c200a1036e56aa0eecc7; sid_ucp_v1=1.0.0-KDljMDcxM2Y1M2RhYzc5NjlmODAzZmM1NWM3NzQ4MmUzMGZlMzQwNGQKDwjR3-qG-QIQqfbojAYYGBoCbGYiIDNiMjU3NTJkNDUwM2MyMDBhMTAzNmU1NmFhMGVlY2M3; ssid_ucp_v1=1.0.0-KDljMDcxM2Y1M2RhYzc5NjlmODAzZmM1NWM3NzQ4MmUzMGZlMzQwNGQKDwjR3-qG-QIQqfbojAYYGBoCbGYiIDNiMjU3NTJkNDUwM2MyMDBhMTAzNmU1NmFhMGVlY2M3; sid_guard=3b25752d4503c200a1036e56aa0eecc7%7C1637497641%7C3023999%7CSun%2C+26-Dec-2021+12%3A27%3A20+GMT; tt_webid=7030387321962808839; MONITOR_WEB_ID=101214498769; _tea_utm_cache_24={%22utm_medium%22:%22wap_search%22}; tt_anti_token=H6JYa4lN2y-74273c4d24d09e7669a6d1952097e10113cbc66cf60efeb0a4ed0e3d8ef2dbfd; ttwid=1%7C5W8_B2qsvafBsxeDQVKBKWESkD1GSnxAEp7GU0qXi3g%7C1637498536%7C6b01a27ae6e0425025042f7cee896e15a8efb1f42ac578f1e87e069df35af7fb; tt_scid=kyqOpBo4VORGKgv9wN82yT.t.bwbhrTQJKG9V5mkWCHmvCoDVx5qGN9zJPsZdvUm9801',
            'referer': 'https://www.toutiao.com/a7032951744313164295/?traffic_source=&in_ogs=&utm_source=&source=search_tab&utm_medium=wap_search&original_source=&in_tfs=&channel=',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.8 Safari/537.36'
        }

        resp = requests.get(url, headers = headers)
        ic(resp.json())

        if resp.status_code == requests.codes.ok:

            json_data = resp.json()['data']

            for item in json_data:

                # 用户名称
                user = item['comment']['user_name']

                # 评论内容
                text = item['comment']['text']

                # 贴子回复数
                reply = item['comment']['reply_count']

                # 评论时间
                times = item['comment']['create_time']
                print(times)
                rls_time = time.strftime('%Y-%m-%d %H:%M', time.localtime(times))

                # 评论点赞数
                stars = item['comment']['digg_count']

                wb.cell(row=count, column=1, value=user)
                wb.cell(row=count, column=2, value=stars)
                wb.cell(row=count, column=3, value=rls_time)
                wb.cell(row=count, column=4, value=reply)
                wb.cell(row=count, column=5, value=text)

                count += 1
                ic(user, stars, rls_time, reply, text)
    # 保存数据
    ws.save('今日头条1.xlsx')
    print('数据保存完毕！')

if __name__ == '__main__':
    spider_page()




