'''
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=66036040659&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=66036040659&score=0&sortType=5&page=1&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=66036040659&score=0&sortType=5&page=2&pageSize=10&isShadowSku=0&rid=0&fold=1
https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=66036040659&score=0&sortType=5&page=3&pageSize=10&isShadowSku=0&rid=0&fold=1
'''

import requests
import openpyxl as op
import time, random
from stylecloud import gen_stylecloud
import pandas as pd
import jieba

save_path = 'jd_评论.xlsx'
def spider_jd():

    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='用户评分')
    wb.cell(row=1, column=2, value='精彩评论')
    space = 2

    for page in range(1, 1000+1):
        print('')
        #获取访问链接
        url = f'https://club.jd.com/comment/productPageComments.action?productId=66036040659&score=0&sortType=5&page={page}&pageSize=10&isShadowSku=0&rid=0&fold=1'
        #添加headers模仿浏览器访问
        headers = {
            'Cookie':'__jdu=1618904846799272696897; shshshfpa=45a0fc5e-e43f-83ec-de31-973b3cf0ae5f-1618904849; shshshfpb=cgzhu4O%2FK8ujiCkcnAquIUg%3D%3D; PCSYCityID=CN_610000_610100_0; areaId=1; ipLoc-djd=1-2800-55811-0; jwotest_product=99; unpl=V2_ZzNtbRVQRBUhABVTKxFZUWILQFURX0sdIF0WU3kZWANvB0VUclRCFnUUR1BnGVgUZwAZXEpcQR1FCEdkeBBVAWMDE1VGZxBFLV0CFSNGF1wjU00zQwBBQHcJFF0uSgwDYgcaDhFTQEJ2XBVQL0oMDDdRFAhyZ0AVRQhHZHseXAxvAhdUR1BFHX0LQFd8EVsMbwUibUVncyV2DkFdch9sBFcCIh8WC0EQcwtPVTYZWwVuCxNYS1JEE30ARVJ4HlQCbgsUbUNnQA%3d%3d; __jdv=76161171|baidu-pinzhuan|t_288551095_baidupinzhuan|cpc|0f3d30c8dba7459bb52f2eb5eba8ac7d_0_f661e9b6a94e49c9b999dda6315795f8|1621051781231; shshshfp=8bb8c20e016f689b32b5f92e968b77d0; __jda=122270672.1618904846799272696897.1618904846.1621047243.1621051781.3; __jdc=122270672; 3AB9D23F7A4B3C9B=AWRX6QXF4LBEQRDZVHR2KKLK4QSDRHWGNNN773BOZCXBCMVPOPAAOQSW2LY7GDN3U42A5KLCJNQFN3LSIS3HXYQAGI; JSESSIONID=9CBE7B47AFE58E9A9014B92F1ABE9F00.s1',
            'Referer':'https://item.jd.com/',
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4503.5 Safari/537.36'
        }
        #获取响应
        resp = requests.get(url, headers = headers).json()
        #获取所有的评论
        comments = resp['comments']
        count = 1

        time.sleep(random.random()*3)
        for comment in comments:
            wb.cell(row=space, column=1, value=comment['score'])
            wb.cell(row=space, column=2, value=comment['content'])
            space += 1
            print(f'第{page}页的第{count}条评论:', comment['score'], comment['content'])
            count += 1
        ws.save(save_path)
        return save_path

#绘制词云
def data_visual():

    rcv_data = pd.read_excel('jd_评论.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    c_title = exist_col['精彩评论'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-cat',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   custom_stopwords=['的', '了', '是啊', '就', '我', '这个', '也', '很', '是', '有', '用', '真的', '做', '都', '和', '还']
                   )
    print('绘图成功！')


if __name__ == '__main__':
    data_visual()


