'''
http://you.163.com/xhr/comment/listByItemByTag.json?__timestamp=1624856130970&itemId=3987228&tag=%E5%85%A8%E9%83%A8&size=20&page=1&orderBy=0&oldItemTag=%E5%85%A8%E9%83%A8&oldItemOrderBy=0&tagChanged=0
http://you.163.com/xhr/comment/listByItemByTag.json?__timestamp=1624856339147&itemId=3987228&tag=%E5%85%A8%E9%83%A8&size=20&page=2&orderBy=0&oldItemTag=%E5%85%A8%E9%83%A8&oldItemOrderBy=0&tagChanged=0
http://you.163.com/xhr/comment/listByItemByTag.json?__timestamp=1624856354947&itemId=3987228&tag=%E5%85%A8%E9%83%A8&size=20&page=3&orderBy=0&oldItemTag=%E5%85%A8%E9%83%A8&oldItemOrderBy=0&tagChanged=0
'''
import requests
import time, random
import pandas as pd
import openpyxl as op
import openpyxl as op
import pandas as pd
import jieba
import json
import re
from pyecharts.faker import Faker
from stylecloud import gen_stylecloud
from collections import Counter
from pyecharts.charts import Bar, Grid, Line, Liquid, Page, Pie, Funnel
from pyecharts.commons.utils import JsCode
from icecream import ic
from pyecharts import options as opts

def spider_page():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='颜色')
    wb.cell(row=1, column=2, value='尺码')
    wb.cell(row=1, column=3, value='评论时间')
    wb.cell(row=1, column=4, value='会员等级')
    wb.cell(row=1, column=5, value='评论点赞')
    wb.cell(row=1, column=6, value='评论内容')
    count = 2

    for page in range(1, 100+1):
        print(f'------------------------正在采集第{page}页数据------------------------')
        url = f'http://you.163.com/xhr/comment/listByItemByTag.json?__timestamp=1636785180888&itemId=3532002&tag=%E5%85%A8%E9%83%A8&size=20&page={page}&orderBy=0&oldItemTag=%E5%85%A8%E9%83%A8&oldItemOrderBy=0&tagChanged=0'
        headers = {
            'Cookie': 'yx_from=web_search_baidu; yx_aui=ada226e7-929f-419c-af99-53ad3eda94f0; mail_psc_fingerprint=01caf6305f28d3e4b8cfe162559acaac; yx_s_device=92db99a-a0c8-22cd-47c3-61d5b24664; yx_but_id=c18807c330874f4aaae2799cd51cdf9fd04f970cabedddc6_v1_nl; P_INFO=18392144506|1636766426|1|yanxuan_web|00&99|null&null&null#CN&null#10#0|&0||18392144506; yx_login_type=0; yx_user_login=true; yx_search_history=%5B%22%u68C9%u8884%u5973%22%2C%22%u7761%u8863%22%2C%22%u68C9%u8884%22%2C%22%u6587%u80F8%22%5D; _ntes_nuid=8b972d5bdb6dba81fd57ccfbac593c87; BAIDU_SSP_lcr=https://www.baidu.com/link?url=zyQS1ZWw9NX5Xw50muitOHx0kkWJG3WT51-8azp0ZDa&wd=&eqid=ca2a415f0000e3c900000005618f20af; _ntes_nnid=f1c2812145357d2b883902c65c421256,1636769976319; yx_delete_cookie_flag=true; yx_stat_seesionId=ada226e7-929f-419c-af99-53ad3eda94f01636769983423; yx_stat_ypmList=; yx_show_painted_egg_shell=false; yx_new_user_modal_show=1; yx_page_key_list=http%3A//you.163.com/search%3Fkeyword%3D%25E6%25A3%2589%25E8%25A2%2584%25E5%25A5%25B3%26timestamp%3D1636769989980%26_stat_search%3Dhistory%26searchWordSource%3D5%26_stat_referer%3Dindex%23page%3D1%26sortType%3D0%26descSorted%3Dtrue%26categoryId%3D0%26matchType%3D0%2Chttp%3A//you.163.com/item/detail%3Fid%3D3991647%26_stat_area%3D1%26_stat_referer%3Dsearch%26_stat_query%3D%25E6%25A3%2589%25E8%25A2%2584%25E5%25A5%25B3%26_stat_count%3D169%26_stat_searchversion%3Dmmoe_model-1.1.0-1.3; yx_stat_seqList=v_315469b8cb%7Cv_f72eac7e53%3B-1%3Bv_0e93fce746%3Bc_6b9da68e5d%3Bv_315469b8cb%3B-1',
            'Referer': 'http://you.163.com/item/detail?id=3987228&_stat_area=3&_stat_referer=search&_stat_query=%E6%96%87%E8%83%B8&_stat_count=132&_stat_searchversion=dcn_model-1.1.0-1.3',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.9 Safari/537.36'
        }

        resp = requests.get(url, headers = headers)

        if resp.status_code == 200:
            comts_List = resp.json()['data']['commentList']

            for item in comts_List:
                # 颜色
                colors = item['skuInfo'][0]

                # 尺码
                size = item['skuInfo'][1]

                # 评论时间
                times = item['createTime']
                content_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(times/1000))

                # 会员等级
                memberLevel = item['memberLevel']

                # 评论点赞
                stars = item['star']

                # 评论内容
                content = item['content']

                print(colors, size, content_time, memberLevel, stars, content)

                wb.cell(row=count, column=1, value=colors)
                wb.cell(row=count, column=2, value=size)
                wb.cell(row=count, column=3, value=content_time)
                wb.cell(row=count, column=4, value=memberLevel)
                wb.cell(row=count, column=5, value=stars)
                wb.cell(row=count, column=6, value=content)

                count += 1
            time.sleep(random.random() * 5)
            ws.save('网易严选大衣1.xlsx')

pd.set_option('display.max_columns', None)  # 显示完整的列
pd.set_option('display.max_rows', None)  # 显示完整的行
pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据

# 读取数据
rcv_data = pd.read_excel('网易严选大衣.xlsx')

# 去除'颜色:'字样
rcv_data.loc[:, '颜色1'] = rcv_data['颜色'].str.replace('颜色:', '')
# 去除'尺码:'字样
rcv_data.loc[:, '尺码1'] = rcv_data['尺码'].str.replace('尺码:', '')

# 存储数据
rcv_data.to_excel('网易严选大衣.xlsx', index=False)

# 删除重复记录和缺失值
rcv_data = rcv_data.drop_duplicates()
rcv_data = rcv_data.dropna()

# 抽样展示
#print(rcv_data.sample(5))

# 获取列内容
c_title = rcv_data['评论内容'].tolist()
# 观影评论词云图
wordlist = jieba.cut(''.join(c_title))
result = ' '.join(wordlist)
stop_words = ['的', '了', '很', '穿', '也', '大', '给']

# 词频设置
all_words = [word for word in result.split(' ') if len(word) > 1 and word not in stop_words]
wordcount = Counter(all_words).most_common(10)

x1_data, y1_data = list(zip(*wordcount))
print(x1_data)
print(y1_data)


def title():
    c = (
        Pie(init_opts=opts.InitOpts(chart_id=40, bg_color='#ADD8E6'))
        .set_global_opts(
            title_opts=opts.TitleOpts(title="----------网易大衣词频展示图----------",
                                      title_textstyle_opts=opts.TextStyleOpts(font_size=36, color='#000000'),
                                      pos_left='center',
                                      pos_top='middle'))
    )
    return c


def funnnel_color():
    c = (
        Funnel(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=41, bg_color='#ADD8E6'))
            .add(
            "词频",
            [list(z) for z in zip(x1_data, y1_data)],
            sort_="ascending",
            label_opts=opts.LabelOpts(position="inside"),
        )
            .set_global_opts(title_opts=opts.TitleOpts(title=""))
    )
    return c

def pie_color():
    v = x1_data
    c = (
        Pie(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=42, bg_color='#ADD8E6'))
            .add(
            "",
            [list(z) for z in zip(v, y1_data)],
            radius=["30%", "75%"],
            center=["25%", "50%"],
            rosetype="radius",
            label_opts=opts.LabelOpts(is_show=False),
        )
            .set_global_opts(title_opts=opts.TitleOpts(title=""))
    )
    return c


def bar_color():
    c = (
        Bar(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=43, bg_color='#ADD8E6'))
            .add_xaxis(x1_data)
            .add_yaxis("词频", y1_data, category_gap=0, color=Faker.rand_color())
            .set_global_opts(title_opts=opts.TitleOpts(title=""))
    )
    return c

def line_color():
    c = (
        Line(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=44, bg_color='#ADD8E6'))
            .add_xaxis(x1_data)
            .add_yaxis(
            "词频",
            y1_data,
            markline_opts=opts.MarkLineOpts(data=[opts.MarkLineItem(type_="average")]),
        )
            .set_global_opts(title_opts=opts.TitleOpts(title=""))
    )
    return c

def page_color():
    page = Page(layout=Page.DraggablePageLayout, page_title="")
    page.add(
        title(),
        funnnel_color(),
        pie_color(),
        bar_color(),
        line_color()
    )
    page.render("大衣.html")
    page.save_resize_html('大衣.html', cfg_file='chart_config.json', dest='词频展示图.html')


def ciyun_visual():

    c_title = rcv_data['评论内容'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img1.jpg'
    gen_stylecloud(text=result,
                   icon_name='fab fa-windows',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   )
    print('绘图成功！')



def visual_ciyun():

    c_title = rcv_data['评论内容'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-cannabis',
                   font_path='msyh.ttc',
                   background_color='white',
                   custom_stopwords=stop_words,
                   output_name=pic,
                   )
    print('绘图成功！')

if __name__ == '__main__':
    #spider_page()
    page_color()
    #visual_ciyun()