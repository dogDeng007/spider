import requests
from icecream import ic
from lxml import etree
import openpyxl as op
import pandas as pd
from pyecharts import options as opts
from pyecharts.charts import Liquid
from pyecharts.commons.utils import JsCode
from fake_useragent import UserAgent


def spider_bage():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='排名')
    wb.cell(row=1, column=2, value='作者')
    wb.cell(row=1, column=3, value='综合得分')
    wb.cell(row=1, column=4, value='视频标题')
    wb.cell(row=1, column=5, value='视频链接')
    wb.cell(row=1, column=6, value='播放数量')
    wb.cell(row=1, column=7, value='弹幕数量')
    wb.cell(row=1, column=8, value='作者详情')
    count = 2

    url = 'https://www.bilibili.com/v/popular/rank/all'

    headers = {
        "cookie": "_uuid=7D3DFA6C-6EB1-F72A-632B-C9AF9B9AD4C627183infoc; buvid3=D25672DE-BD2D-4E7C-B79E-DB356316D023167639infoc; sid=aylq5kgg; fingerprint=84acc3579a53d0eba78d769e71574df6; buvid_fp=BA184AFC-F4DC-408A-8897-D0EDEA653CE5148812infoc; buvid_fp_plain=BA184AFC-F4DC-408A-8897-D0EDEA653CE5148812infoc; DedeUserID=434541726; DedeUserID__ckMd5=448fda6ab5098e5e; SESSDATA=78a505c8%2C1643594982%2Cdfa35*81; bili_jct=1d9f4e960fb0ae7fe1de53663029874b; bsource=search_baidu; CURRENT_FNVAL=80; blackside_state=1; rpdid=|(u)YJR~R~)m0J'uYk)ku)~~)",
        "referer": "https://www.bilibili.com/",
        #"user-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.8 Safari/537.36"
        "user-agent": str(UserAgent().random)
    }

    resp = requests.get(url, headers = headers, timeout=15)

    if resp.status_code == requests.codes.ok:
        html_data = etree.HTML(resp.text)

        # 获取榜单所有列表
        lis = html_data.xpath("//div[@class='rank-list-wrap']/ul[@class='rank-list']/li[@class='rank-item']")

        for li in lis:

            # 排名
            sort = li.xpath("./div[@class='num']/text()")
            sort = ''.join(sort)

            # 作者
            author = li.xpath("./div[@class='content']/div[@class='info']/div[@class='detail']/a/span[@class='data-box up-name']/text()")
            author = ''.join(author).strip()

            # 综合得分
            score = li.xpath("./div[@class='content']/div[@class='info']/div[@class='pts']/div/text()")
            score = ''.join(score)

            # 视频标题
            title = li.xpath("./div[@class='content']/div[@class='info']/a[@class='title']/text()")
            title = ''.join(title)

            # 视频链接
            links = li.xpath("./div[@class='content']/div[@class='img']/a/@href")
            links = ''.join(links).strip()[2:]

            # 播放数量
            video_num = li.xpath("./div[@class='content']/div[@class='info']/div[@class='detail']/span[@class='data-box'][1]/text()")
            video_num = ''.join(video_num).strip()

            # 弹幕数量
            barrage_num = li.xpath("./div[@class='content']/div[@class='info']/div[@class='detail']/span[@class='data-box'][2]/text()")
            barrage_num = ''.join(barrage_num).strip()

            # 作者详情
            detail_auth = li.xpath(".//div[@class='content']/div[@class='info']/div[@class='detail']/a/@href")
            detail_auth = ['https:' + i for i in detail_auth]
            detail_auth = ''.join(detail_auth)

            wb.cell(row=count, column=1, value=sort)
            wb.cell(row=count, column=2, value=author)
            wb.cell(row=count, column=3, value=score)
            wb.cell(row=count, column=4, value=title)
            wb.cell(row=count, column=5, value=links)
            wb.cell(row=count, column=6, value=video_num)
            wb.cell(row=count, column=7, value=barrage_num)
            wb.cell(row=count, column=8, value=detail_auth)
            ic(sort, author, score, title,links, video_num, barrage_num, detail_auth)
            count += 1

    ws.save('哔哩哔哩Top100.xlsx')

def anay_data():
    # 读取数据
    df = pd.read_excel('哔哩哔哩Top100.xlsx')
    # 删除空格
    pd_data = df.dropna(subset=['播放数量', '弹幕数量'])

    # 格式化数据播放数量
    # 去除’万‘
    pd_data['播放数量'] = pd_data['播放数量'].str.replace('万', '')

    # 转换格式 万->10000
    pd_data['弹幕数量'] = pd_data['弹幕数量'].map(lambda x: float(x[:-1]) * 10000 if ('万' in x) else float(x))

    # 处理后的数据另存为
    pd_data.to_excel('哔哩哔哩Top101.xlsx')

def find_data():
    # 读取数据
    rcv_data = pd.read_excel('哔哩哔哩Top101.xlsx')

    pd.set_option('display.max_columns', None)  # 显示完整的列
    pd.set_option('display.max_rows', None)  # 显示完整的行
    pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据

    # 最多播放
    max_video_num = rcv_data[rcv_data['播放数量'] == rcv_data['播放数量'].max()]
    ic(max_video_num)

    # 最多弹幕
    max_bag_num = rcv_data[rcv_data['弹幕数量'] == rcv_data['弹幕数量'].max()]
    ic(max_bag_num)


def visual_data():
    # 读取数据
    rcv_data = pd.read_excel('哔哩哔哩Top101.xlsx')

    # 平均数
    mean_score = rcv_data['综合得分'].mean()

    # 最大值
    max_score = rcv_data['综合得分'].max()
    ic(mean_score/max_score)


c = (
    Liquid()
    .add(
        "lq",
        [0.4027],
        label_opts=opts.LabelOpts(
            font_size=50,
            formatter=JsCode(
                """function (param) {
                    return (Math.floor(param.value * 10000) / 100) + '%';
                }"""
            ),
            position="inside",
        ),
    )
    .set_global_opts(title_opts=opts.TitleOpts(title="综合得分平均数-水滴图"))
    .render("综合得分平均值水滴图.shtml")
)


if __name__ == '__main__':
    #spider_bage()
    #anay_data()
    # find_data()
    visual_data()