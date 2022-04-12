import requests
import re
import openpyxl as op
import pandas as pd
from stylecloud import gen_stylecloud
import jieba


def spider_wb():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='用户id')
    wb.cell(row=1, column=2, value='作者名称')
    wb.cell(row=1, column=3, value='作者座右铭')
    wb.cell(row=1, column=4, value='发帖时间')
    wb.cell(row=1, column=5, value='发帖内容')

    count = 2

    # 爬取第一页
    print(f'----------正在打印第1页数据----------')
    url = 'https://m.weibo.cn/comments/hotflow?id=4665088419561949&mid=4665088419561949&max_id_type=0'
    print('当前url是：', url)

    headers = {
        'cookie': 'SUB=_2A25NyTOqDeRhGeVG7lAZ9S_PwjiIHXVvMl3irDV6PUJbktB-LVDmkW1NT7e8qozwK1pqWVKX_PsKk5dhdCyPXwW1; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WFGibRIp_iSfMUfmcr5kb295NHD95Q01h-E1h-pe0.XWs4DqcjLi--fi-2Xi-2Ni--fi-z7iKysi--Ri-8si-zXi--fi-88i-zce7tt; _T_WM=98961943286; MLOGIN=1; WEIBOCN_FROM=1110006030; XSRF-TOKEN=70a1e0; M_WEIBOCN_PARAMS=oid%3D4648381753067388%26luicode%3D20000061%26lfid%3D4648381753067388%26uicode%3D20000061%26fid%3D4648381753067388',
        'referer': 'Referer: https://m.weibo.cn/detail/4665088419561949',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4573.0 Safari/537.36'
    }

    resp = requests.get(url, headers=headers).json()
    wb_info = resp['data']['data']

    for item in wb_info:
        user_id = item.get('user')['id']  # 用户id

        author = item['user']['screen_name']  # 作者名称

        auth_sign = item['user']['description']  # 作者座右铭

        time = str(item['created_at']).split(' ')[1:4]
        rls_time = '-'.join(time)  # 发帖时间

        text = ''.join(re.findall('[\u4e00-\u9fa5]', item['text']))  # 发帖内容

        wb.cell(row=count, column=1, value=user_id)
        wb.cell(row=count, column=2, value=author)
        wb.cell(row=count, column=3, value=auth_sign)
        wb.cell(row=count, column=4, value=rls_time)
        wb.cell(row=count, column=5, value=text)
        print(user_id, author, auth_sign, rls_time, text)
        count += 1

    # 爬取第一页往后
    max_id = resp['data']['max_id']
    count = 17
    for page in range(2, 50 + 1):
        print(f'----------正在打印第{page}页数据----------')
        if page > 16:
            max_id_type = 1
        else:
            max_id_type = 0
        url = f'https://m.weibo.cn/comments/hotflow?id=4665088419561949&mid=4665088419561949&max_id={max_id}&max_id_type={max_id_type}'
        print('当前url是：', url)

        resp = requests.get(url, headers=headers)

        if resp.status_code == 200:
            comments = resp.json()

            if comments['ok'] == 1:
                max_id = comments['data']['max_id']
                wb_info = comments['data']['data']

                for item in wb_info:
                    user_id = item.get('user')['id']  # 用户id

                    author = item['user']['screen_name']  # 作者名称

                    auth_sign = item['user']['description']  # 作者座右铭

                    time = str(item['created_at']).split(' ')[1:4]
                    rls_time = '-'.join(time)  # 发帖时间

                    text = ''.join(re.findall('[\u4e00-\u9fa5]', item['text']))  # 发帖内容


                    wb.cell(row=count, column=1, value=user_id)
                    wb.cell(row=count, column=2, value=author)
                    wb.cell(row=count, column=3, value=auth_sign)
                    wb.cell(row=count, column=4, value=rls_time)
                    wb.cell(row=count, column=5, value=text)
                    count += 1
                    print(user_id, author, auth_sign, rls_time, text)

                page += 1
    ws.save('微博疫情1.xlsx')


def data_visual():

    rcv_data = pd.read_excel('./微博疫情1.xlsx')
    exist_col = rcv_data.dropna()  # 删除空行
    c_title = exist_col['发帖内容'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    print(result)
    pic = 'img1.jpg'
    gen_stylecloud(text=result,
                   icon_name='fas fa-mask',
                   font_path='msyh.ttc',
                   background_color='black',
                   output_name=pic,
                   custom_stopwords=['你', '我', '的', '了', '在', '吧', '是', '也', '都', '不', '吗', '就', '人', '去', '有', '还', '没', '被', '好', '呢']
                   )
    print('绘图成功！')

if __name__ == '__main__':
    #spider_wb()
    data_visual()