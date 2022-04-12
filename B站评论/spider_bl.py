'''
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17205690584633020348_1636423610452&jsonp=jsonp&next=0&type=1&oid=336587753&mode=3&plat=1&_=1636423611589
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17205690584633020348_1636423610453&jsonp=jsonp&next=2&type=1&oid=336587753&mode=3&plat=1&_=1636424178396
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17205690584633020348_1636423610454&jsonp=jsonp&next=3&type=1&oid=336587753&mode=3&plat=1&_=1636424183583
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17205690584633020348_1636423610455&jsonp=jsonp&next=4&type=1&oid=336587753&mode=3&plat=1&_=1636424187787

https://api.bilibili.com/x/v2/reply/main?callback=jQuery17209449843921008754_1645669390753&jsonp=jsonp&next=0&type=1&oid=207910661&mode=3&plat=1&_=1645669397186
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17209449843921008754_1645669390754&jsonp=jsonp&next=2&type=1&oid=207910661&mode=3&plat=1&_=1645669402606
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17209449843921008754_1645669390755&jsonp=jsonp&next=3&type=1&oid=207910661&mode=3&plat=1&_=1645669405742
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17209449843921008754_1645669390756&jsonp=jsonp&next=4&type=1&oid=207910661&mode=3&plat=1&_=1645669434682
https://api.bilibili.com/x/v2/reply/main?callback=jQuery17209449843921008754_1645669390757&jsonp=jsonp&next=5&type=1&oid=207910661&mode=3&plat=1&_=1645669444541
'''
import time
import requests
import openpyxl as op
import pandas as pd
import jieba
import json
import re
from pyecharts.faker import Faker
from stylecloud import gen_stylecloud
from collections import Counter
from pyecharts.charts import Bar, Grid, Line, Liquid, Page, Pie, Funnel
from pyecharts.commons.utils import JsCode
from icecream import ic
from pyecharts import options as opts

def spider_page():
    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='评论者')
    wb.cell(row=1, column=2, value='性别')
    wb.cell(row=1, column=3, value='评论时间')
    wb.cell(row=1, column=4, value='点赞人数')
    wb.cell(row=1, column=5, value='评论内容')
    wb.cell(row=1, column=6, value='会员等级')
    count = 2

    for page in range(1, 100 + 1):
        print(f'-----------------正在爬取第{page}页数据-----------------')
        time_thick = int(time.time() * 1000)
        #url = f'https://api.bilibili.com/x/v2/reply/main?callback=jQuery17205690584633020348_{1636423610452 + page}&jsonp=jsonp&next={page}&type=1&oid=336587753&mode=3&plat=1&_={time_thick}'
        url = f'https://api.bilibili.com/x/v2/reply/main?callback=jQuery17209449843921008754_{1645669390753 + page}&jsonp=jsonp&next={page}&type=1&oid=207910661&mode=3&plat=1&_={time_thick}'

        headers = {
            "cookie": "_uuid=7C8813FE-61109-D9C6-4B13-D310E32E12AC753205infoc; buvid3=A5304D2E-E953-4E8F-B868-1ECA3C0C032F167629infoc; b_nut=1641296253; PVID=1; LIVE_BUVID=AUTO3116447133906552; buvid4=BC4E610D-2F51-F596-98CA-CF663F84989291013-022021308-y7FzQHcGAvocfhaWGr9Lbg%3D%3D; i-wanna-go-back=-1; CURRENT_BLACKGAP=0; sid=lduetpqq; fingerprint=c49148099004e31ec3883a8a67ccdce9; buvid_fp_plain=undefined; DedeUserID=434541726; DedeUserID__ckMd5=448fda6ab5098e5e; SESSDATA=05b6725f%2C1660957098%2C3b454*21; bili_jct=e0b55882ac885cd69b1e0ba45b487f78; b_ut=5; buvid_fp=8b5d7c44f3824f3559e451fde3d35acc; blackside_state=1; rpdid=|(u)YJR)YRR~0J'uYRlR|lu~J; CURRENT_FNVAL=4048; b_lsid=2CCED2C8_17F292289F1; bsource=share_source_weixinchat; bp_video_offset_434541726=630601353480110100",
            'referer': 'https://message.bilibili.com/',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4878.0 Safari/537.36'
        }

        resp = requests.get(url, headers=headers)

        if resp.status_code == requests.codes.ok:
            # 获取resp响应
            text = resp.text[41:-1]

            # 转换json格式
            json_data = json.loads(text)

            # 获取所有评论
            datas = json_data['data']['replies']
            if datas != None:
                for item in datas:
                    # 评论者
                    name = item['member']['uname']

                    # 性别
                    sex = item['member']['sex']

                    # 评论时间
                    ctime = item.get('ctime')
                    content_time = time.strftime('%Y-%m-%d %H:%M', time.localtime(ctime))

                    # 点赞人数
                    star = item['like']

                    # 评论内容
                    cmts = item['content']['message']
                    text = ''.join(re.findall('[\u4e00-\u9fa5 0-9?？。.,，]', cmts))  # 去除多余符号

                    # 会员等级
                    member_level = item['member']['level_info']['current_level']

                    wb.cell(row=count, column=1, value=name)
                    wb.cell(row=count, column=2, value=sex)
                    wb.cell(row=count, column=3, value=content_time)
                    wb.cell(row=count, column=4, value=star)
                    wb.cell(row=count, column=5, value=text)
                    wb.cell(row=count, column=6, value=member_level)
                    count += 1

                    ic(name, sex, content_time, star, cmts, member_level)

    ws.save('哔哩哔哩1.xlsx')

pd.set_option('display.max_columns', None)  # 显示完整的列
pd.set_option('display.max_rows', None)  # 显示完整的行
pd.set_option('display.expand_frame_repr', False)  # 设置不折叠数据

# 读取数据
rcv_data = pd.read_excel('哔哩哔哩1.xlsx')

# 删除重复记录和缺失值
rcv_data = rcv_data.drop_duplicates()
rcv_data = rcv_data.dropna()

# 抽样展示
print(rcv_data.sample(5))

# 获取列内容
c_title = rcv_data['评论内容'].tolist()
# 观影评论词云图
wordlist = jieba.cut(''.join(c_title))
result = ' '.join(wordlist)
stop_words = ['是', '的', '系列', '热词', '大', '了', '抽', '一个', '会员', '我们', '我', '但', '也', '吧', '人', '不', '在', '打', '人不骗']


# 词频设置
all_words = [word for word in result.split(' ') if len(word) > 1 and word not in stop_words]
wordcount = Counter(all_words).most_common(10)

x1_data, y1_data = list(zip(*wordcount))
print(x1_data)
print(y1_data)

def title():
    c = (
        Pie(init_opts=opts.InitOpts(chart_id=40, bg_color='#ADD8E6'))
        .set_global_opts(
            title_opts=opts.TitleOpts(title="----------EDG词频图----------",
                                      title_textstyle_opts=opts.TextStyleOpts(font_size=36, color='#000000'),
                                      pos_left='center',
                                      pos_top='middle'))
    )
    return c


def funnnel_color():
    c = (
        Funnel(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=41, bg_color='#ADD8E6'))
            .add(
            "词频",
            [list(z) for z in zip(x1_data, y1_data)],
            sort_="ascending",
            label_opts=opts.LabelOpts(position="inside"),
        )
            .set_global_opts(title_opts=opts.TitleOpts(title="词频"))
    )
    return c

def pie_color():
    v = x1_data
    c = (
        Pie(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=42, bg_color='#ADD8E6'))
            .add(
            "",
            [list(z) for z in zip(v, y1_data)],
            radius=["30%", "75%"],
            center=["25%", "50%"],
            rosetype="radius",
            label_opts=opts.LabelOpts(is_show=False),
        )
            .set_global_opts(title_opts=opts.TitleOpts(title="EDG词频"))
    )
    return c


def bar_color():
    c = (
        Bar(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=43, bg_color='#ADD8E6'))
            .add_xaxis(x1_data)
            .add_yaxis("词频", y1_data, category_gap=0, color=Faker.rand_color())
            .set_global_opts(title_opts=opts.TitleOpts(title="EDG"))
    )
    return c

def line_color():
    c = (
        Line(init_opts=opts.InitOpts(width="1600px", height="1000px",chart_id=44, bg_color='#ADD8E6'))
            .add_xaxis(x1_data)
            .add_yaxis(
            "词频",
            y1_data,
            markline_opts=opts.MarkLineOpts(data=[opts.MarkLineItem(type_="average")]),
        )
            .set_global_opts(title_opts=opts.TitleOpts(title="EDG"))
    )
    return c

def page_color():
    page = Page(layout=Page.DraggablePageLayout, page_title="EDG")
    page.add(
        title(),
        funnnel_color(),
        pie_color(),
        bar_color(),
        line_color()
    )
    page.render("edg.html")
    page.save_resize_html('edg.html', cfg_file='chart_config.json', dest='EDG词频展示图.html')


def ciyun_visual():

    c_title = rcv_data['评论内容'].tolist()
    # 观影评论词云图
    wordlist = jieba.cut(''.join(c_title))
    result = ' '.join(wordlist)
    pic = 'img1.jpg'
    gen_stylecloud(text=result,
                   icon_name='fab fa-windows',
                   font_path='msyh.ttc',
                   background_color='white',
                   output_name=pic,
                   )
    print('绘图成功！')


if __name__ == '__main__':
    #spider_page()
    ciyun_visual()
    #cipin_visual()
    page_color()