'''
斗鱼妹子
颜值：https://www.douyu.com/gapi/rknc/directory/yzRec/2
     https://www.douyu.com/gapi/rknc/directory/yzRec/4
交友：https://www.douyu.com/gapi/rkc/directory/mixList/2_1555/1
     https://www.douyu.com/gapi/rkc/directory/mixList/2_1555/2
     https://www.douyu.com/gapi/rkc/directory/mixList/2_1555/3
舞蹈：https://www.douyu.com/gapi/rkc/directory/mixList/2_1008/1
     https://www.douyu.com/gapi/rkc/directory/mixList/2_1008/2
'''
import requests
import openpyxl as op
from icecream import ic

def spider_douyu():

    ws = op.Workbook()
    wb = ws.create_sheet(index=0)

    wb.cell(row=1, column=1, value='照片链接')
    wb.cell(row=1, column=2, value='房间标题')
    wb.cell(row=1, column=3, value='房间名称')
    wb.cell(row=1, column=4, value='主播分类')
    count = 2

    save_pic = []
    for page in range(1, 5+1):
        print(f'--------正在爬取第{page}页--------')
        url = f'https://www.douyu.com/gapi/rknc/directory/yzRec/{page}'
        headers = {
            'cookie':'dy_did=00d862396f61af86adf2c44300091601; acf_did=00d862396f61af86adf2c44300091601; acf_auth=5e4cEm%2BWNxXtumjXb%2BL2omhygmIKsSJq%2FnGXtSSCldwD58XT%2Buu4YqWYwz5oQNbv9dUwivRoVJQZzjBIlXpLDxB0h8J%2BOeyUbMHoDnSSFGeVBKM109Tm6s4; dy_auth=ac54KMZB159aRLDhzgZSf5P5LIzzKVZKLTfa0O0dPIvuPqyY54CboGkC60ppxS%2BY6TKiT%2FSsgRUHzAUlIh3cG6nevsPatEzrPA1WpzNQ6ObpuzCIuJKWMXA; wan_auth37wan=2b8ac21c432cV8yreMDDyT%2BLR%2FVAmoIocsjN8yPJWkIPd6wRzkzR1GdXNcjg9%2FVO8M9UJzpVidRtobuPmbYheDS4gGQLVUOEARWBDTywMdSt7wEmBlY; acf_uid=139162360; acf_username=139162360; acf_nickname=%E5%B8%95%E5%8D%9A%E9%9B%B7%E5%85%8B666; acf_own_room=0; acf_groupid=1; acf_phonestatus=1; acf_ct=0; acf_ltkid=10823032; acf_biz=1; acf_stk=a90a348ca88dec4f; acf_avatar=//apic.douyucdn.cn/upload/avanew/face/201705/07/20/3d9f7186451336f7b97992c2fba06bf8_; Hm_lvt_e99aee90ec1b2106afe7ec3b199020a7=1645955426,1645955640,1645955714,1646094095; Hm_lpvt_e99aee90ec1b2106afe7ec3b199020a7=1646094122',
            'referer':'https://www.douyu.com/g_yz',
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4484.7 Safari/537.36'
        }
        #获取浏览器返回的JSON结果
        resp = requests.get(url,headers = headers).json()
        #获取每一页的120个颜值小姐姐信息
        girls = resp['data']['rl']

        #数据存入Excel
        for girl in girls:
            pic_links = girl['rs1']     # 所有照片链接
            room_title = girl['rn']     # 房间标题
            room_name = girl['nn']      # 房间名称
            category = girl['c2name']   # 主播分类
            save_pic.append((pic_links, room_name))

            wb.cell(row=count, column=1, value=pic_links)
            wb.cell(row=count, column=2, value=room_title)
            wb.cell(row=count, column=3, value=room_name)
            wb.cell(row=count, column=4, value=category)
            count += 1
    ws.save('斗鱼颜值1.xlsx')
    return save_pic

def down_dy(save_pic):
    count = 0
    for num, (pic_img, pic_name) in enumerate(save_pic):
        r = requests.get(pic_img)
        pic = r.content

        try:
            with open('./pictures1/{}.jpg'.format(pic_name), 'wb') as fin:
                print(f'正在爬取第{count}张图片')
                fin.write(pic)
                print('{}.jpg----下载成功'.format(pic_name))
        except:
            print('下载失败！')
        count += 1

if __name__ == '__main__':
    down_dy(save_pic=spider_douyu())



